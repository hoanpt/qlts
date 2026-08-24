import sqlite3
import openpyxl
import os
import re
import uuid
import time
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = r'f:\QLTS\server\prisma\dev.db'

conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

now_ms = int(time.time() * 1000)
now_iso = time.strftime('%Y-%m-%dT%H:%M:%S.000Z', time.gmtime())

DEPT_CODE_MAP = {
    'PKDK': 'PKDK', 'PKĐK': 'PKDK', 'PHÒNG KHÁM': 'PKDK', 'PHONG KHAM': 'PKDK',
    'XN': 'XN', 'XÉT NGHIỆM': 'XN', 'XET NGHIEM': 'XN',
    'BNN': 'BNN', 'BỆNH NGHỀ NGHIỆP': 'BNN', 'BENH NGHE NGHIEP': 'BNN',
    'DD': 'DD', 'DINH DƯỠNG': 'DD', 'DINH DUONG': 'DD', 'DINHDUONG': 'DD',
    'SKMT': 'SKMT', 'SỨC KHỔE MÔI TRƯỜNG': 'SKMT', 'SKMTYTTH': 'SKMT',
    'SKSS': 'SKSS', 'SỨC KHỎE SINH SẢN': 'SKSS',
    'KHNV': 'KHNV', 'KẾ HOẠCH NGHIỆP VỤ': 'KHNV',
    'TTGDSK': 'TTGDSK', 'TRUYỀN THÔNG': 'TTGDSK',
    'KSTCT': 'KSTCT', 'KÝ SINH TRÙNG': 'KSTCT', 'KST': 'KSTCT',
    'PCBKLN': 'PCBKLN', 'BKLN': 'PCBKLN', 'BỆNH KHÔNG LÂY NHIỄM': 'PCBKLN',
    'DVTYT': 'DVTYT', 'DƯỢC': 'DVTYT', 'DUOC': 'DVTYT', 'VTYT': 'DVTYT',
    'HIV': 'HIV', 'PHÒNG CHỐNG HIV': 'HIV',
    'PCBTN': 'PCBTN', 'BTN': 'PCBTN', 'BỆNH TRUYỀN NHIỄM': 'PCBTN',
    'TCKT': 'TCKT', 'TÀI CHÍNH KẾ TOÁN': 'TCKT', 'TAI CHINH': 'TCKT',
    'TCHC': 'TCHC', 'TỔ CHỨC HÀNH CHÍNH': 'TCHC', 'TO CHUC': 'TCHC', 'TC-HC': 'TCHC',
    'KDYTQT': 'KDYTQT', 'KIỂM DỊCH': 'KDYTQT', 'KIEM DICH': 'KDYTQT'
}

def get_dept_id(name_or_code):
    norm = str(name_or_code or '').upper().strip()
    target_code = 'TCHC'
    for k, v in DEPT_CODE_MAP.items():
        if k in norm:
            target_code = v
            break
    cursor.execute("SELECT id FROM Department WHERE code = ?", (target_code,))
    row = cursor.fetchone()
    return row[0] if row else 1

# 1. Clear CNTT assets only
print("1. Rebuilding CNTT assets as unified consolidated computer units (Màn hình + CPU)...")
cursor.execute("DELETE FROM Asset WHERE managingUnit = 'CNTT'")
conn.commit()

cntt_dir = r'f:\QLTS\TS\CNTT'
cntt_files = [
    ('1. PKĐK 2026 chốt.xlsx', '2026 PKDK'),
    ('2. XN KK2026 chốt.xlsx', '2026 XN phan hoi'),
    ('3. BNN KK 2026..xlsx', '2026 BNN'),
    ('4. Dinh dưỡng KK2026 chốt.xlsx', 'Ktoan mới'),
    ('5. SKMT -KK-2026.xlsx', '2026 TB CNTT'),
    ('6. SKSS CNTT KK2026.xlsx', '2023 SKSS'),
    ('7. KHNV KK2026.xlsx', '2026 KHNV'),
    ('8. TTGDSK KK2026 chốt.xlsx', '2026 TTGDSK'),
    ('9. KSTCT KK2026.xlsx', '2026 KST.CT'),
    ('10. BKLN- KK 2026 phan hồi.xlsx', '2026'),
    ('11. Dược KK 2026.xlsx', '2026 Dược'),
    ('12. HIV 2026.xlsx', 'CNTT'),
    ('13. BTN-CNTT-kinh phí phan hôi.xlsx', 'BTN cơ sở 2'),
    ('14. TCKT 2026 chốt.xlsx', '2026 TCKT'),
    ('15. TCHC chốt.xlsx', '2026 TCHC'),
    ('16. KDYTQT chốt.xlsx', '2024 Kiểm dịch')
]

seen_codes = set()
cursor.execute("SELECT assetCode FROM Asset")
for r in cursor.fetchall():
    seen_codes.add(r[0])

cntt_pc_count = 0
cntt_lap_count = 0
cntt_print_count = 0
cntt_other_count = 0

def insert_cntt_asset(code, name, dept_id, year=None, price=None, specs='', note=''):
    global cntt_pc_count, cntt_lap_count, cntt_print_count, cntt_other_count
    code = str(code).strip()
    name = str(name).strip()
    if not code or not name or len(name) < 2:
        return

    # clean bad names
    if any(k in name.lower() for k in ['thành viên', 'ông.', 'bà.', 'trần văn vũ', 'phan thanh hoàn', 'trần thị liên', 'huỳnh thị thanh tú']):
        return

    final_code = code
    idx = 1
    while final_code in seen_codes:
        idx += 1
        final_code = f"{code}_{idx}"
    seen_codes.add(final_code)

    qr = str(uuid.uuid4())[:12]
    cursor.execute("""
        INSERT INTO Asset (
            assetCode, name, categoryId, departmentId, managingUnit, location,
            assignedTo, yearInUse, originalPrice, status,
            bookQuantity, actualQuantity, quantityDifference, specifications, note,
            qrCode, createdAt, updatedAt
        ) VALUES (?, ?, 2, ?, 'CNTT', 'Cơ sở 1', ?, ?, ?, 'DANG_SU_DUNG', 1, 1, 0, ?, ?, ?, ?, ?)
    """, (
        final_code, name, dept_id, None, year, price, specs or None, note or None, qr, now_ms, now_ms
    ))

for fname, s_name in cntt_files:
    p = os.path.join(cntt_dir, fname)
    if not os.path.exists(p):
        continue
    dept_id = get_dept_id(fname)
    dept_code = fname.split('.')[1].strip().split(' ')[0]

    try:
        wb = openpyxl.load_workbook(p, data_only=True)
        if s_name not in wb.sheetnames:
            s_name = wb.sheetnames[-1]
        ws = wb[s_name]

        header_r = 15
        for r in range(1, min(ws.max_row + 1, 22)):
            v = str(ws.cell(r, 2).value or '').lower()
            if 'tên' in v or 'tài sản' in v or 'stt' in str(ws.cell(r, 1).value or '').lower():
                header_r = r
                break

        current_pc = None

        for r in range(header_r + 1, ws.max_row + 1):
            c1 = ws.cell(r, 1).value
            c2 = str(ws.cell(r, 2).value or '').strip()
            c3 = ws.cell(r, 3).value
            c4 = str(ws.cell(r, 4).value or '').strip()
            c5 = str(ws.cell(r, 5).value or '').strip()

            if not c2 or any(c2.startswith(k) for k in ['Mẫu', 'BIÊN', 'Thành viên', 'Trần Văn', 'Phan Thanh', 'Trần Thị', 'Huỳnh Thị', '1.', '2.', '3.', '4.', '5.']):
                continue

            # Check if main numbered row
            is_num = False
            if c1 is not None:
                try:
                    int(float(str(c1).strip()))
                    is_num = True
                except:
                    pass

            if is_num:
                # Save pending PC
                if current_pc:
                    # Construct consolidated PC
                    screen_info = f"Màn hình: {current_pc['screen_code']}" if current_pc['screen_code'] else ''
                    cpu_info = f"CPU: {current_pc['cpu_code']}" if current_pc['cpu_code'] else ''
                    specs_parts = [p for p in [screen_info, cpu_info] if p]
                    specs_text = " | ".join(specs_parts)
                    
                    pc_code = current_pc['cpu_code'] or current_pc['screen_code'] or f"PC/{dept_code}-{current_pc['stt']:03d}"
                    if not pc_code.startswith('PC/') and not pc_code.startswith('CPU/'):
                        pc_code = f"PC/{pc_code}"

                    pc_name = current_pc['name']
                    if 'máy vi tính' not in pc_name.lower() and 'máy tính' not in pc_name.lower():
                        pc_name = f"Bộ máy vi tính {pc_name}"

                    year_v = current_pc['year'] or (int(float(current_pc['screen_year'])) if current_pc['screen_year'] and current_pc['screen_year'].isdigit() else None)
                    insert_cntt_asset(pc_code, pc_name, dept_id, year=year_v, price=current_pc['price'], specs=specs_text)
                    cntt_pc_count += 1
                    current_pc = None

                # Check device type
                c2_lower = c2.lower()
                if any(k in c2_lower for k in ['máy vi tính', 'bộ máy', 'máy tính để bàn', 'pc ']) or c2_lower == 'máy tính':
                    current_pc = {
                        'stt': int(float(str(c1))),
                        'name': c2,
                        'screen_code': '',
                        'cpu_code': '',
                        'screen_year': '',
                        'cpu_year': '',
                        'year': None,
                        'price': ws.cell(r, 8).value or ws.cell(r, 7).value or ws.cell(r, 6).value
                    }
                else:
                    # Standalone Laptop, Printer, Scanner, Network
                    code_val = c5 if c5 and not c5.isdigit() and len(c5) > 2 else (c4 if c4 and len(c4) > 2 else f"CNTT-{dept_code}-{r:03d}")
                    if code_val.lower() in ['stt', 'tài sản', 'none', 'null']:
                        code_val = f"CNTT-{dept_code}-{r:03d}"
                    
                    year_v = None
                    try:
                        for y_col in [3, 4, 5]:
                            y = ws.cell(r, y_col).value
                            if y and 1990 <= int(float(y)) <= 2030:
                                year_v = int(float(y))
                                break
                    except:
                        pass

                    price_v = None
                    for p_col in [8, 7, 6, 9]:
                        pv = ws.cell(r, p_col).value
                        if isinstance(pv, (int, float)) and pv > 0:
                            price_v = float(pv)
                            break

                    insert_cntt_asset(code_val, c2, dept_id, year=year_v, price=price_v)
                    if any(k in c2_lower for k in ['laptop', 'xách tay']):
                        cntt_lap_count += 1
                    elif any(k in c2_lower for k in ['in', 'scan', 'quét', 'photo']):
                        cntt_print_count += 1
                    else:
                        cntt_other_count += 1
            else:
                # Sub-row (+ Màn hình, + Khối CPU)
                if current_pc:
                    c2_lower = c2.lower()
                    code_val = c5 if c5 and not c5.isdigit() and len(c5) > 2 else (c4 if c4 and len(c4) > 2 else '')
                    year_v = str(c3 or '')
                    if 'màn hình' in c2_lower:
                        current_pc['screen_code'] = code_val
                        current_pc['screen_year'] = year_v
                    elif 'cpu' in c2_lower or 'khối' in c2_lower or 'thùng' in c2_lower:
                        current_pc['cpu_code'] = code_val
                        current_pc['cpu_year'] = year_v

        if current_pc:
            screen_info = f"Màn hình: {current_pc['screen_code']}" if current_pc['screen_code'] else ''
            cpu_info = f"CPU: {current_pc['cpu_code']}" if current_pc['cpu_code'] else ''
            specs_parts = [p for p in [screen_info, cpu_info] if p]
            specs_text = " | ".join(specs_parts)
            pc_code = current_pc['cpu_code'] or current_pc['screen_code'] or f"PC/{dept_code}-{current_pc['stt']:03d}"
            if not pc_code.startswith('PC/') and not pc_code.startswith('CPU/'):
                pc_code = f"PC/{pc_code}"
            insert_cntt_asset(pc_code, current_pc['name'], dept_id, price=current_pc['price'], specs=specs_text)
            cntt_pc_count += 1

    except Exception as e:
        print(f"Error in {fname}: {e}")

print(f"\n✓ Hoàn tất cấu trúc Khối CNTT theo Bộ máy vi tính hoàn chỉnh:")
print(f"  • Bộ máy vi tính (Màn hình + CPU): {cntt_pc_count}")
print(f"  • Laptop: {cntt_lap_count}")
print(f"  • Máy in & Scan: {cntt_print_count}")
print(f"  • Thiết bị mạng & khác: {cntt_other_count}")
print(f"  => Tổng cộng CNTT: {cntt_pc_count + cntt_lap_count + cntt_print_count + cntt_other_count} tài sản")

# 2. Fix Maintenance Requests data (Clean out any mismatch like Ổn áp -> CNTT)
print("\n2. Fixing Maintenance Requests and sample cases for 100% accurate department & unit mapping...")
cursor.execute("DELETE FROM MaintenanceRequest")

# Find real representative assets
# PC in PKDK
cursor.execute("SELECT id, name FROM Asset WHERE managingUnit = 'CNTT' AND departmentId = 1 AND name LIKE '%máy%' LIMIT 1")
pc_pkdk = cursor.fetchone() or (1, 'Bộ máy vi tính Dell')

# Printer in XN
cursor.execute("SELECT id, name FROM Asset WHERE managingUnit = 'CNTT' AND departmentId = 2 AND (name LIKE '%in%' OR name LIKE '%canon%') LIMIT 1")
print_xn = cursor.fetchone() or (2, 'Máy in Canon LBP 2900')

# Medical Device in XN
cursor.execute("SELECT id, name FROM Asset WHERE managingUnit = 'DUOC' AND departmentId = 2 LIMIT 1")
med_xn = cursor.fetchone() or (500, 'Máy sinh hóa tự động')

# Stabilizer / Electrical in TCHC (Ổn áp Lioa / Tủ điện thuộc TCHC)
cursor.execute("SELECT id, name FROM Asset WHERE managingUnit = 'TCHC' AND (name LIKE '%ổn áp%' OR name LIKE '%điện%' OR name LIKE '%quạt%') LIMIT 1")
on_ap_tchc = cursor.fetchone() or (1200, 'Máy ổn áp Lioa 10kVA')

# Insert correct maintenance cases
maintenance_cases = [
    (
        pc_pkdk[0], 1, 'BS. Trương Tấn Nam', '0913987654',
        'Bộ máy vi tính phòng khám không lên nguồn sau sự cố điện, quạt nguồn không quay.',
        'HIGH', 'COMPLETED', 350000, 'Công ty Tin Học CDC',
        'Đã thay thế bộ nguồn Huntkey 450W, vệ sinh thùng máy, máy hoạt động ổn định.',
        'KS. Phan Thanh Hoàn', 'CNTT', 'Phòng Khám Đa Khoa (Tầng 1)',
        '2026-01-12T08:30:00.000Z', '2026-01-12T16:00:00.000Z'
    ),
    (
        print_xn[0], 2, 'KS. Nguyễn Trường Duy', '0905123456',
        'Máy in kết quả xét nghiệm bị kẹt giấy liên tục, bản in bị vệt đen dọc mép giấy.',
        'MEDIUM', 'COMPLETED', 220000, 'Dịch vụ Thiết bị Văn phòng',
        'Đã thay thế rulo cuốn giấy và gạt từ hộp mực.',
        'KS. Huỳnh Bá Thành', 'CNTT', 'Phòng Xét nghiệm Hóa sinh (Tầng 2)',
        '2026-01-20T09:00:00.000Z', '2026-01-20T14:30:00.000Z'
    ),
    (
        med_xn[0], 2, 'DS. Trần Thị Ngọc Diệp', '0905888999',
        'Máy xét nghiệm báo lỗi nhiệt độ buồng ủ phản ứng không đạt chuẩn 37°C.',
        'URGENT', 'COMPLETED', 1800000, 'Hãng Thiết Bị Y Tế Việt Nam',
        'Kỹ sư hãng đã hiệu chỉnh cảm biến nhiệt độ và thay bộ gia nhiệt buồng ủ.',
        'KS. Phạm Phú Ân', 'DUOC', 'Khu vực Xét nghiệm Chuyên sâu (Tầng 3)',
        '2026-02-05T10:15:00.000Z', '2026-02-07T15:00:00.000Z'
    ),
    (
        on_ap_tchc[0], 15, 'Ông. Trần Liên', '0905333444',
        'Ổn áp tổng cấp nguồn khu vực máy chủ và phòng họp bị sụt áp, phát tiếng kêu rè rè.',
        'HIGH', 'IN_PROGRESS', None, None,
        'Thợ điện phòng TCHC đang kiểm tra chổi than và rơ le tự ngắt của ổn áp.',
        'Ông. Lê Xuân Lộc', 'TCHC', 'Phòng Kỹ thuật Điện TCHC (Tầng 1)',
        '2026-02-21T08:00:00.000Z', None
    )
]

for c in maintenance_cases:
    cursor.execute("""
        INSERT INTO MaintenanceRequest (
            assetId, departmentId, requestedBy, contactPhone, issueDescription,
            priority, status, repairCost, repairVendor, repairNote,
            technicianName, managingUnit, locationDetail, requestDate, completedDate,
            createdAt, updatedAt
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (
        c[0], c[1], c[2], c[3], c[4],
        c[5], c[6], c[7], c[8], c[9],
        c[10], c[11], c[12], c[13], c[14],
        now_iso, now_iso
    ))

conn.commit()

# Print summary
cursor.execute("SELECT managingUnit, COUNT(*) FROM Asset GROUP BY managingUnit")
print("\n=======================================================")
print("TỔNG HỢP SỐ LIỆU TÀI SẢN TOÀN TRUNG TÂM SAU KHI ĐỒNG BỘ:")
print("=======================================================")
for r in cursor.fetchall():
    print(f"  • Khối Quản Lý {r[0]}: {r[1]} tài sản")

cursor.execute("SELECT COUNT(*) FROM Asset")
print(f"\n=> TỔNG CỘNG TÀI SẢN TOÀN HỆ THỐNG: {cursor.fetchone()[0]} tài sản")

conn.close()
