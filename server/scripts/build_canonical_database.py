import sqlite3
import openpyxl
import os
import re
import uuid
import time
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = r'f:\QLTS\server\prisma\dev.db'

# Department mapping
DEPT_CODE_MAP = {
    'PKDK': 'PKDK', 'PKĐK': 'PKDK', 'PHÒNG KHÁM': 'PKDK', 'PHONG KHAM': 'PKDK',
    'XN': 'XN', 'XÉT NGHIỆM': 'XN', 'XET NGHIEM': 'XN',
    'BNN': 'BNN', 'BỆNH NGHỀ NGHIỆP': 'BNN', 'BENH NGHE NGHIEP': 'BNN',
    'DD': 'DD', 'DINH DƯỠNG': 'DD', 'DINH DUONG': 'DD', 'DINHDUONG': 'DD',
    'SKMT': 'SKMT', 'SỨC KHỎE MÔI TRƯỜNG': 'SKMT', 'SKMTYTTH': 'SKMT',
    'SKSS': 'SKSS', 'SỨC KHỎE SINH SẢN': 'SKSS',
    'KHNV': 'KHNV', 'KẾ HOẠCH NGHIỆP VỤ': 'KHNV',
    'TTGDSK': 'TTGDSK', 'TRUYỀN THÔNG': 'TTGDSK',
    'KSTCT': 'KSTCT', 'KÝ SINH TRÙNG': 'KSTCT', 'KST': 'KSTCT',
    'PCBKLN': 'PCBKLN', 'BKLN': 'PCBKLN', 'BỆNH KHÔNG LÂY NHIỄM': 'PCBKLN',
    'DVTYT': 'DVTYT', 'DƯỢC': 'DVTYT', 'DUOC': 'DVTYT', 'VTYT': 'DVTYT',
    'HIV': 'HIV', 'PHÒNG CHỐNG HIV': 'HIV',
    'PCBTN': 'PCBTN', 'BTN': 'PCBTN', 'BỆNH TRUYỀN NHIỄM': 'PCBTN',
    'TCKT': 'TCKT', 'TÀI CHÍNH KẾ TOÁN': 'TCKT', 'TAI CHINH': 'TCKT',
    'TCHC': 'TCHC', 'TỔ CHỨC HÀNH CHÍNH': 'TCHC', 'TO CHUC': 'TCHC', 'TC-HC': 'TCHC',
    'KDYTQT': 'KDYTQT', 'KIỂM DỊCH': 'KDYTQT', 'KIEM DICH': 'KDYTQT'
}

def get_dept_id(cursor, name_or_code):
    norm = str(name_or_code or '').upper().strip()
    target_code = 'TCHC'
    for k, v in DEPT_CODE_MAP.items():
        if k in norm:
            target_code = v
            break
    cursor.execute("SELECT id FROM Department WHERE code = ?", (target_code,))
    row = cursor.fetchone()
    return row[0] if row else 1

def build_canonical_database():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    now_ms = int(time.time() * 1000)
    now_iso = time.strftime('%Y-%m-%dT%H:%M:%S.000Z', time.gmtime())

    # =========================================================================
    # 1. SETUP CATEGORIES
    # =========================================================================
    cursor.execute("DELETE FROM AssetCategory")
    categories = [
        (1, 'DUOC', 'Trang thiết bị Y tế (Khoa Dược quản lý)', 'Máy xét nghiệm, siêu âm, X-quang, chẩn đoán hình ảnh, bảo quản vắc xin, tủ sấy, kính hiển vi...'),
        (2, 'CNTT', 'Thiết bị Công nghệ thông tin (Tổ CNTT quản lý)', 'Máy tính (PC), Laptop, Máy in, Máy Scan, Thiết bị mạng, Server, Màn hình...'),
        (3, 'TCHC', 'Thiết bị Hành chính & CCDC (Phòng TCHC quản lý)', 'Bàn làm việc, ghế xoay, tủ sắt, tủ gỗ, bàn họp, quạt, giường inox...'),
        (4, 'TBVP_TOANHA', 'Cơ sở vật chất & Hạ tầng tòa nhà theo tầng (TCHC)', 'Công tắc điện, ổ cắm, đèn led, tủ điện, PCCC, máy bơm... theo tầng')
    ]
    for c in categories:
        cursor.execute("INSERT INTO AssetCategory (id, code, name, description) VALUES (?, ?, ?, ?)", c)
    conn.commit()

    # =========================================================================
    # 2. SETUP 16 DEPARTMENTS
    # =========================================================================
    cursor.execute("DELETE FROM Department")
    depts = [
        (1, 'PKDK', 'Phòng Khám Đa Khoa', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (2, 'XN', 'Khoa Xét Nghiệm - CĐHA - TDCN', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (3, 'BNN', 'Khoa Bệnh Nghề Nghiệp', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (4, 'DD', 'Khoa Dinh Dưỡng', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (5, 'SKMT', 'Khoa Sức Khỏe Môi Trường - YTTH', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (6, 'SKSS', 'Khoa Sức Khỏe Sinh Sản', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (7, 'KHNV', 'Phòng Kế Hoạch Nghiệp Vụ', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (8, 'TTGDSK', 'Khoa Truyền Thông Giáo Dục Sức Khỏe', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (9, 'KSTCT', 'Khoa Ký Sinh Trùng - Côn Trùng', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (10, 'PCBKLN', 'Khoa Phòng Chống Bệnh Không Lây Nhiễm', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (11, 'DVTYT', 'Khoa Dược - Vật Tư Y Tế', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (12, 'HIV', 'Khoa HIV/AIDS và QLĐTNC', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (13, 'PCBTN', 'Khoa Phòng Chống Bệnh Truyền Nhiễm', 'Cơ sở 2', 'Bàn Thạch, Hòa Vang, Đà Nẵng'),
        (14, 'TCKT', 'Phòng Tài Chính - Kế Toán', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (15, 'TCHC', 'Phòng Tổ Chức - Hành Chính', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng'),
        (16, 'KDYTQT', 'Khoa Kiểm Dịch Y Tế Quốc Tế', 'Cơ sở 1', '118 Lê Đình Lý, Thanh Khê, Đà Nẵng')
    ]
    for d in depts:
        cursor.execute("INSERT INTO Department (id, code, name, location, description, createdAt, updatedAt) VALUES (?, ?, ?, ?, ?, ?, ?)",
                       (d[0], d[1], d[2], d[3], d[4], now_iso, now_iso))
    conn.commit()

    # Clear Asset table
    cursor.execute("DELETE FROM Asset")
    conn.commit()
    print("Asset table cleared. Building canonical 3,684 dataset...\n")

    seen_codes = set()
    total_inserted = 0

    def insert_asset(code, name, cat_id, dept_id, managing_unit, location='Cơ sở 1',
                     loc_detail='', assigned_to='', year=None, price=None, status='DANG_SU_DUNG',
                     book_qty=1, actual_qty=1, qty_diff=0, source='', note='',
                     floor=None, building_asset=0, specs=''):
        nonlocal total_inserted
        code = str(code).strip()
        name = str(name).strip()
        if not code or not name or len(name) < 2:
            return False

        final_code = code
        dup_idx = 1
        while final_code in seen_codes:
            dup_idx += 1
            final_code = f"{code}_{dup_idx}"
        seen_codes.add(final_code)

        qr_code = str(uuid.uuid4())[:12]
        
        try:
            cursor.execute("""
                INSERT INTO Asset (
                    assetCode, name, categoryId, departmentId, managingUnit, location, locationDetail,
                    assignedTo, yearInUse, originalPrice, currentValue, status,
                    bookQuantity, actualQuantity, quantityDifference, source, note, specifications,
                    floor, buildingAsset, qrCode, createdAt, updatedAt
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                final_code, name, cat_id, dept_id, managing_unit, location, loc_detail or None,
                assigned_to or None, year, price, None, status,
                book_qty, actual_qty, qty_diff, source or None, note or None, specs or None,
                floor, building_asset, qr_code, now_ms, now_ms
            ))
            total_inserted += 1
            return True
        except Exception as e:
            print(f"Error inserting {final_code}: {e}")
            return False

    # =========================================================================
    # PART 1: KHỐI KHOA DƯỢC - TRANG THIẾT BỊ Y TẾ (1,497 tài sản)
    # File: TỔNG HỢP TS CCDC TBYT CDC 2026.xlsx (13 sheets)
    # =========================================================================
    main_file = r'f:\QLTS\TS\TỔNG HỢP TS CCDC TBYT CDC 2026.xlsx'
    print(f"1. Importing Khối Khoa Dược (TBYT) from {os.path.basename(main_file)}...")
    wb_main = openpyxl.load_workbook(main_file, data_only=True)
    tbyt_count = 0

    for s in wb_main.sheetnames:
        ws = wb_main[s]
        dept_id = get_dept_id(cursor, s)
        dept_loc = 'Cơ sở 2' if 'BTN' in s.upper() or 'PCBTN' in s.upper() else 'Cơ sở 1'
        
        header_r = 14
        for r in range(1, min(ws.max_row + 1, 20)):
            v1 = str(ws.cell(r, 1).value or '').lower()
            v2 = str(ws.cell(r, 2).value or '').lower()
            if 'stt' in v1 or 'tài sản' in v2:
                header_r = r
                break

        current_sub_loc = None
        s_count = 0
        for r in range(header_r + 2, ws.max_row + 1):
            col1 = ws.cell(r, 1).value
            col2 = ws.cell(r, 2).value
            col3 = ws.cell(r, 3).value
            col4 = ws.cell(r, 4).value
            col5 = ws.cell(r, 5).value
            col6 = ws.cell(r, 6).value
            col7 = ws.cell(r, 7).value
            col8 = ws.cell(r, 8).value
            col9 = ws.cell(r, 9).value
            col10 = ws.cell(r, 10).value
            col11 = ws.cell(r, 11).value
            col12 = ws.cell(r, 12).value
            col13 = ws.cell(r, 13).value
            col14 = ws.cell(r, 14).value

            if col1 is None and col2 and not col3:
                v2_str = str(col2).strip()
                if not v2_str.startswith('Mẫu') and not v2_str.startswith('BIÊN'):
                    current_sub_loc = v2_str
                continue

            if not col2:
                continue

            name = str(col2).strip()
            if len(name) < 2 or name.lower().startswith('mẫu số') or name.lower().startswith('biên bản'):
                continue

            # Original asset code (Col 3)
            code = str(col3 or '').strip()
            if not code or code.lower() in ['none', 'null', '-', '']:
                code = f"TS-{s}-{s_count + 1:03d}"

            # Year
            year_val = None
            try:
                if col4 and 1990 <= int(float(col4)) <= 2030:
                    year_val = int(float(col4))
            except:
                pass

            # Quantities
            book_q = int(float(col5)) if col5 and isinstance(col5, (int, float)) else 1
            act_q = int(float(col6)) if col6 and isinstance(col6, (int, float)) else book_q
            diff_q = int(float(col7)) if col7 and isinstance(col7, (int, float)) else (act_q - book_q)

            # Price
            price_val = None
            try:
                if col8 and isinstance(col8, (int, float)) and col8 > 0:
                    price_val = float(col8)
                elif col9 and isinstance(col9, (int, float)) and col9 > 0:
                    price_val = float(col9) / (act_q or 1)
            except:
                pass

            # Status
            stat_str = str(col12 or '').lower()
            status = 'DANG_SU_DUNG'
            if 'hỏng' in stat_str or 'đề nghị sửa' in stat_str: status = 'HONG'
            elif 'thanh lý' in stat_str: status = 'CHO_THANH_LY'
            elif 'không sử dụng' in stat_str: status = 'KHONG_SU_DUNG'

            loc_det = str(col11 or current_sub_loc or '').strip()
            note_det = str(col14 or '').strip()
            source_det = str(col10 or '').strip()

            if insert_asset(
                code=code, name=name, cat_id=1, dept_id=dept_id, managing_unit='DUOC',
                location=dept_loc, loc_detail=loc_det, assigned_to=loc_det, year=year_val,
                price=price_val, status=status, book_qty=book_q, actual_qty=act_q, qty_diff=diff_q,
                source=source_det, note=note_det
            ):
                s_count += 1
                tbyt_count += 1

        print(f"  ✓ Sheet '{s}': {s_count} assets (TSPK..., TSXN...)")

    print(f"  => TỔNG CỘNG KHỐI KHOA DƯỢC (TBYT): {tbyt_count} tài sản\n")

    # =========================================================================
    # PART 2: KHỐI TỔ CNTT (896 thiết bị)
    # Files: 16 CNTT chốt files
    # =========================================================================
    print("2. Importing Khối Tổ CNTT from 16 files (f:\\QLTS\\TS\\CNTT\\)...")
    cntt_dir = r'f:\QLTS\TS\CNTT'
    cntt_count = 0

    CNTT_TARGETS = [
        ('1. PKĐK 2026 chốt.xlsx', '2026 PKDK'),
        ('2. XN KK2026 chốt.xlsx', '2026 XN phan hoi'),
        ('3. BNN KK 2026..xlsx', '2026 BNN'),
        ('4. Dinh dưỡng KK2026 chốt.xlsx', 'Ktoan mới'),
        ('5. SKMT -KK-2026.xlsx', '2026 TB CNTT'),
        ('6. SKSS CNTT KK2026.xlsx', '2023 SKSS'),
        ('7. KHNV KK2026.xlsx', '2026 KHNV'),
        ('8. TTGDSK KK2026 chốt.xlsx', '2026 TTGDSK'),
        ('9. KSTCT KK2026.xlsx', '2026 KST.CT'),
        ('10. BKLN- KK 2026 phan hồi.xlsx', '2026'),
        ('11. Dược KK 2026.xlsx', '2026 Dược'),
        ('12. HIV 2026.xlsx', 'CNTT'),
        ('13. BTN-CNTT-kinh phí phan hôi.xlsx', 'BTN cơ sở 2'),
        ('14. TCKT 2026 chốt.xlsx', '2026 TCKT'),
        ('15. TCHC chốt.xlsx', '2026 TCHC'),
        ('16. KDYTQT chốt.xlsx', '2024 Kiểm dịch')
    ]

    for fname, sheet_name in CNTT_TARGETS:
        p = os.path.join(cntt_dir, fname)
        if not os.path.exists(p):
            continue
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[-1]
            ws = wb[sheet_name]
            dept_id = get_dept_id(cursor, fname)
            f_count = 0

            header_r = 15
            code_col = 3
            code_new_col = 4
            for r in range(1, min(ws.max_row + 1, 20)):
                for c in range(1, min(ws.max_column + 1, 10)):
                    val = str(ws.cell(r, c).value or '').lower()
                    if 'mã' in val and 'mới' in val:
                        code_new_col = c
                        header_r = r
                    elif 'mã' in val:
                        code_col = c
                        header_r = r
                if header_r != 15:
                    break

            parent_machine = ''
            for r in range(header_r + 1, ws.max_row + 1):
                col1 = ws.cell(r, 1).value
                col2 = ws.cell(r, 2).value
                col_old = ws.cell(r, code_col).value if code_col else None
                col_new = ws.cell(r, code_new_col).value if code_new_col else None

                if not col2 and not col1:
                    continue

                col2_str = str(col2 or '').strip()
                if not col2_str or col2_str.startswith('Mẫu') or col2_str.startswith('BIÊN') or col2_str.startswith('THIẾT BỊ'):
                    continue

                if col1 is not None and str(col1).strip().isdigit() and not col_old and not col_new:
                    parent_machine = col2_str
                    continue

                exact_code = str(col_new or col_old or '').strip()
                if not exact_code or exact_code.lower() in ['none', 'null', '-', '']:
                    if parent_machine:
                        exact_code = f"CNTT-{fname[:4].strip()}-{r:03d}"
                    else:
                        continue

                # Filter out junk header rows
                if exact_code.lower() in ['stt', 's\nt\nt', 'tài sản', 'mã số', 'mã số mới', 'a', 'b', 'c', 'd'] or exact_code.isdigit():
                    continue

                full_name = col2_str
                if parent_machine and ('màn hình' in col2_str.lower() or 'cpu' in col2_str.lower()):
                    full_name = f"{parent_machine} ({col2_str.replace('+', '').strip()})"

                year_val = None
                try:
                    for y_col in [5, 4, 6]:
                        y_c = ws.cell(r, y_col).value
                        if y_c and 1990 <= int(float(y_c)) <= 2030:
                            year_val = int(float(y_c))
                            break
                except:
                    pass

                price_val = None
                for p_col in [8, 7, 6, 9]:
                    p_c = ws.cell(r, p_col).value
                    if isinstance(p_c, (int, float)) and p_c > 0:
                        price_val = float(p_c)
                        break

                if insert_asset(
                    code=exact_code, name=full_name, cat_id=2, dept_id=dept_id,
                    managing_unit='CNTT', location='Cơ sở 1', loc_detail=f"Phòng {fname[:10]}",
                    year=year_val, price=price_val, status='DANG_SU_DUNG'
                ):
                    f_count += 1
                    cntt_count += 1

            print(f"  ✓ File '{fname}' [{sheet_name}]: {f_count} CNTT assets (MH/..., CPU/..., LAP/..., IN/...)")
        except Exception as e:
            print(f"  Error reading {fname}: {e}")

    print(f"  => TỔNG CỘNG KHỐI TỔ CNTT: {cntt_count} tài sản\n")

    # =========================================================================
    # PART 3: KHỐI PHÒNG TCHC (1,291 tài sản)
    # - 294 Hành chính & CCDC văn phòng (Bàn, ghế, tủ, quạt, giường inox...)
    # - 997 Cơ sở vật chất hạ tầng gắn liền tòa nhà theo 8 tầng
    # File: 15 TCHC TS KK T1 2026.xlsx
    # =========================================================================
    print("3. Importing Khối Phòng TCHC (Hành chính & Hạ tầng Tòa nhà 8 tầng)...")
    tchc_file = r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien\15 TCHC TS KK T1 2026.xlsx'
    tchc_hc_count = 0
    tchc_toanha_count = 0

    if os.path.exists(tchc_file):
        wb_tchc = openpyxl.load_workbook(tchc_file, data_only=True)
        
        # 3.1. Hành chính & CCDC Sheets: 'T1', 'tang 1', 'tang 3', 'tang 4', 't5', 't6', 't7', 'CS2', 'Nha dat'
        # 3.2. Tòa nhà Sheets: 'TB DIEN 2022 TANG HAM', 'TB DIEN 2022 T1', 'TB DIEN 2022 T2 ', 'TB DIEN 2022 T3', 'TB DIEN 2022 T4', 'TB DIEN 2022 T5', 'TB DIEN 2022 t6', 'TB DIEN 2022 t7', 'BAN GIAO'

        for s in wb_tchc.sheetnames:
            ws = wb_tchc[s]
            if ws.max_row <= 2:
                continue

            s_lower = s.lower()
            is_building = 'tb dien' in s_lower or 'ban giao' in s_lower or 'hầm' in s_lower
            
            floor_tag = 'Tầng 1'
            if 't7' in s_lower or 'tầng 7' in s_lower: floor_tag = 'Tầng 7'
            elif 't6' in s_lower or 'tầng 6' in s_lower: floor_tag = 'Tầng 6'
            elif 't5' in s_lower or 'tầng 5' in s_lower: floor_tag = 'Tầng 5'
            elif 't4' in s_lower or 'tầng 4' in s_lower: floor_tag = 'Tầng 4'
            elif 't3' in s_lower or 'tầng 3' in s_lower: floor_tag = 'Tầng 3'
            elif 't2' in s_lower or 'tầng 2' in s_lower: floor_tag = 'Tầng 2'
            elif 'hầm' in s_lower or 'tang ham' in s_lower: floor_tag = 'Tầng Hầm'

            current_room = floor_tag
            sheet_imported = 0

            for r in range(1, ws.max_row + 1):
                col1 = ws.cell(r, 1).value
                col2 = ws.cell(r, 2).value
                col3 = ws.cell(r, 3).value
                col4 = ws.cell(r, 4).value
                col5 = ws.cell(r, 5).value

                # Section/room header
                if col1 in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'A', 'B', 'C', 'D'] or (col1 is None and col2 and str(col2).startswith('P.')):
                    current_room = f"{floor_tag} - {str(col2 or col3 or '').strip()}"
                    continue

                if col1 is not None and str(col1).strip().isdigit():
                    stt = int(col1)
                    name_val = str(col2 or col3 or '').strip()
                    code_val = str(col3 or '').strip() if col3 and col3 != name_val else ''

                    if not name_val or name_val.lower().startswith('tên') or len(name_val) < 2:
                        continue

                    # Quantities
                    qty = 1
                    for q_col in [4, 5, 6]:
                        v = ws.cell(r, q_col).value
                        if isinstance(v, (int, float)) and v > 0:
                            qty = int(v)
                            break

                    if is_building:
                        # Hạ tầng tòa nhà
                        exact_code = f"TN-{floor_tag.replace(' ', '')}-{stt:03d}"
                        cat_id = 4 # TBVP_TOANHA
                        b_asset = 1
                        fl_val = floor_tag
                        if insert_asset(
                            code=exact_code, name=name_val, cat_id=cat_id, dept_id=15,
                            managing_unit='TCHC', location='Cơ sở 1', loc_detail=current_room,
                            book_qty=qty, actual_qty=qty, floor=fl_val, building_asset=b_asset
                        ):
                            sheet_imported += 1
                            tchc_toanha_count += 1
                    else:
                        # Thiết bị hành chính
                        exact_code = code_val if (code_val and len(code_val) > 2 and not code_val.isdigit()) else f"TCHC-HC-{floor_tag.replace(' ', '')}-{stt:03d}"
                        cat_id = 3 # TCHC
                        b_asset = 0
                        fl_val = floor_tag
                        if insert_asset(
                            code=exact_code, name=name_val, cat_id=cat_id, dept_id=15,
                            managing_unit='TCHC', location='Cơ sở 1', loc_detail=current_room,
                            book_qty=qty, actual_qty=qty, floor=fl_val, building_asset=b_asset
                        ):
                            sheet_imported += 1
                            tchc_hc_count += 1

            print(f"  ✓ Sheet '{s}': {sheet_imported} assets (is_building={is_building})")

    print(f"  => TCHC Thiết bị Hành chính: {tchc_hc_count} | TCHC Hạ tầng Tòa nhà: {tchc_toanha_count}")
    print(f"  => TỔNG CỘNG KHỐI PHÒNG TCHC: {tchc_hc_count + tchc_toanha_count} tài sản\n")

    # =========================================================================
    # PART 4: SETUP 3 DISTINCT INVENTORY TEAMS + 16 DEPARTMENT REPS
    # =========================================================================
    print("4. Setting up 3 distinct inventory teams & 16 department reps...")
    cursor.execute("DROP TABLE IF EXISTS InventoryCommitteeMember")
    cursor.execute("""
    CREATE TABLE InventoryCommitteeMember (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fullName TEXT NOT NULL,
        position TEXT NOT NULL,
        role TEXT NOT NULL,
        departmentId INTEGER,
        scope TEXT DEFAULT 'ALL',
        isActive INTEGER DEFAULT 1,
        displayOrder INTEGER DEFAULT 0,
        createdAt DATETIME DEFAULT CURRENT_TIMESTAMP,
        updatedAt DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    """)

    committee = [
        # Chung toàn Hội Đồng
        ('Ông. Nguyễn Đại Vĩnh', 'Giám đốc', 'CHUTICH', None, 'ALL', 1),
        ('Ông. Hồ Phú Quảng', 'Trưởng phòng TC - KT', 'UYVIEN', 14, 'ALL', 2),

        # TỔ 1: Tổ Kiểm Kê Trang Thiết Bị Y Tế (Khoa Dược chủ trì)
        ('Bà. Mai Thị Tính', 'Phụ trách Khoa Dược - VTYT', 'TOTRUONG_TBYT', 11, 'DUOC', 10),
        ('Bà. Trần Thị Ngọc Diệp', 'Dược sĩ Khoa Dược - VTYT', 'THANHVIEN_DUOC', 11, 'DUOC', 11),
        ('Bà. Lê Thị Thanh Thủy', 'Dược sĩ Khoa Dược - VTYT', 'THANHVIEN_DUOC', 11, 'DUOC', 12),
        ('Ông. Phạm Phú Ân', 'Kỹ sư Thiết bị y tế', 'THANHVIEN_DUOC', 11, 'DUOC', 13),

        # TỔ 2: Tổ Kiểm Kê Thiết Bị CNTT (Tổ CNTT / Phòng KHNV chủ trì)
        ('Ông. Trần Văn Vũ', 'Trưởng phòng KHNV', 'TOTRUONG_CNTT', 7, 'CNTT', 20),
        ('Ông. Huỳnh Bá Thành', 'Kỹ sư CNTT', 'THANHVIEN_CNTT', 7, 'CNTT', 21),
        ('Ông. Lê Xuân Lộc', 'Kỹ sư CNTT', 'THANHVIEN_CNTT', 7, 'CNTT', 22),
        ('Ông. Nguyễn Văn Hùng', 'Chuyên viên CNTT', 'THANHVIEN_CNTT', 7, 'CNTT', 23),

        # TỔ 3: Tổ Kiểm Kê Thiết Bị Hành Chính & Hạ Tầng Tòa Nhà (Phòng TCHC chủ trì)
        ('Ông. Trần Liên', 'Trưởng phòng TC - HC', 'TOTRUONG_TCHC', 15, 'TCHC', 30),
        ('Ông. Phạm Phú Ân', 'Cán bộ phòng TCHC', 'THANHVIEN_TCHC', 15, 'TCHC', 31),
        ('Ông. Lê Xuân Lộc', 'Cán bộ phòng TCHC', 'THANHVIEN_TCHC', 15, 'TCHC', 32),
        ('Bà. Lê Thị Thanh Thủy', 'Cán bộ phòng TCHC', 'THANHVIEN_TCHC', 15, 'TCHC', 33),
        ('Ông. Huỳnh Bá Thành', 'Cán bộ phòng TCHC', 'THANHVIEN_TCHC', 15, 'TCHC', 34),

        # ĐẠI DIỆN 16 KHOA / PHÒNG
        ('Ông. Trương Tấn Nam', 'Trưởng Phòng khám đa khoa', 'DAIDIEN_KHOA', 1, 'PKDK', 40),
        ('Ông. Nguyễn Trường Duy', 'Phó Trưởng khoa XN-CĐHA-TDCN', 'DAIDIEN_KHOA', 2, 'XN', 41),
        ('Ông. Dương Ấm Mậu', 'Trưởng khoa Bệnh nghề nghiệp', 'DAIDIEN_KHOA', 3, 'BNN', 42),
        ('Bà. Nguyễn Thị Thu Trang', 'Phó Trưởng khoa Dinh dưỡng', 'DAIDIEN_KHOA', 4, 'DD', 43),
        ('Ông. Lê Văn Cường', 'Phó Trưởng khoa Sức khỏe môi trường - YTTH', 'DAIDIEN_KHOA', 5, 'SKMT', 44),
        ('Bà. Trần Thị Dạ Thảo', 'Trưởng khoa Sức khỏe sinh sản', 'DAIDIEN_KHOA', 6, 'SKSS', 45),
        ('Ông. Trần Văn Vũ', 'Trưởng phòng Kế hoạch nghiệp vụ', 'DAIDIEN_KHOA', 7, 'KHNV', 46),
        ('Bà. Phan Thị Mỹ Lệ', 'Phó Trưởng khoa Truyền thông GDSK', 'DAIDIEN_KHOA', 8, 'TTGDSK', 47),
        ('Ông. Nguyễn Như Tiến', 'Phó Trưởng khoa Ký sinh trùng - Côn trùng', 'DAIDIEN_KHOA', 9, 'KSTCT', 48),
        ('Bà. Bùi Thị Long Cảnh', 'Trưởng khoa Phòng chống Bệnh không lây nhiễm', 'DAIDIEN_KHOA', 10, 'PCBKLN', 49),
        ('Bà. Mai Thị Tính', 'Phụ trách Khoa Dược - VTYT', 'DAIDIEN_KHOA', 11, 'DVTYT', 50),
        ('Ông. Cao Minh Thông', 'Trưởng khoa HIV/AIDS và QLĐTNC', 'DAIDIEN_KHOA', 12, 'HIV', 51),
        ('Ông. Đặng Quang Ánh', 'Phó Trưởng khoa Phòng chống Bệnh truyền nhiễm', 'DAIDIEN_KHOA', 13, 'PCBTN', 52),
        ('Ông. Hồ Phú Quảng', 'Trưởng phòng Tài chính - Kế toán', 'DAIDIEN_KHOA', 14, 'TCKT', 53),
        ('Ông. Trần Liên', 'Trưởng phòng Tổ chức - Hành chính', 'DAIDIEN_KHOA', 15, 'TCHC', 54),
        ('Ông. Phan Văn Bửu', 'Trưởng khoa Kiểm dịch Y tế Quốc tế', 'DAIDIEN_KHOA', 16, 'KDYTQT', 55)
    ]

    for m in committee:
        cursor.execute("""
            INSERT INTO InventoryCommitteeMember (fullName, position, role, departmentId, scope, isActive, displayOrder, createdAt, updatedAt)
            VALUES (?, ?, ?, ?, ?, 1, ?, ?, ?)
        """, (m[0], m[1], m[2], m[3], m[4], m[5], now_iso, now_iso))

    conn.commit()

    # =========================================================================
    # SUMMARY
    # =========================================================================
    print("=======================================================")
    print("KẾT QUẢ TỔNG HỢP DỮ LIỆU CHUẨN XÁC TOÀN HỆ THỐNG:")
    print("=======================================================")
    cursor.execute("SELECT managingUnit, COUNT(*) FROM Asset GROUP BY managingUnit")
    for r in cursor.fetchall():
        print(f"  • Khối Quản Lý {r[0]}: {r[1]} tài sản")

    cursor.execute("SELECT COUNT(*) FROM Asset")
    print(f"\n=> TỔNG CỘNG TÀI SẢN TOÀN TRUNG TÂM: {cursor.fetchone()[0]} tài sản")

    conn.close()

if __name__ == '__main__':
    build_canonical_database()
