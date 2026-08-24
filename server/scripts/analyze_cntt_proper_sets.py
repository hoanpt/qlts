import openpyxl
import os
import re
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

cntt_dir = r'f:\QLTS\TS\CNTT'
files = [f for f in sorted(os.listdir(cntt_dir)) if f.endswith('.xlsx') and not f.startswith('~$')]

total_pc_sets = 0
total_laptops = 0
total_printers = 0
total_others = 0

for fname in files:
    p = os.path.join(cntt_dir, fname)
    wb = openpyxl.load_workbook(p, data_only=True)
    
    # Choose 2026 sheet
    chosen_sheet = wb.sheetnames[0]
    for s in wb.sheetnames:
        if '2026' in s or 'CNTT' in s or 'chốt' in s.lower():
            chosen_sheet = s
            break
    ws = wb[chosen_sheet]

    # Find header row
    header_r = 15
    for r in range(1, min(ws.max_row + 1, 20)):
        v = str(ws.cell(r, 2).value or '').lower()
        if 'tên' in v or 'tài sản' in v:
            header_r = r
            break

    pc_in_file = 0
    laptop_in_file = 0
    printer_in_file = 0
    other_in_file = 0

    current_pc = None
    
    for r in range(header_r + 1, ws.max_row + 1):
        col1 = ws.cell(r, 1).value
        col2 = str(ws.cell(r, 2).value or '').strip()
        col3 = ws.cell(r, 3).value
        col4 = str(ws.cell(r, 4).value or '').strip()
        col5 = str(ws.cell(r, 5).value or '').strip()

        if not col2 or col2.startswith('Mẫu') or col2.startswith('BIÊN') or col2.startswith('Thành viên') or col2.startswith('Trần Văn') or col2.startswith('Phan Thanh') or col2.startswith('Trần Thị') or col2.startswith('Huỳnh Thị'):
            continue

        # If col1 is a number, it's a main item!
        is_num = False
        if col1 is not None:
            try:
                int(float(str(col1).strip()))
                is_num = True
            except:
                pass

        if is_num:
            # Save previous PC if any
            if current_pc:
                pc_in_file += 1
                current_pc = None

            # Check if this item is a PC or Laptop or Printer or Network
            col2_lower = col2.lower()
            if any(k in col2_lower for k in ['máy vi tính', 'bộ máy', 'máy tính để bàn', 'pc ']) or col2_lower == 'máy tính':
                current_pc = {
                    'stt': col1,
                    'name': col2,
                    'screen_code': '',
                    'cpu_code': '',
                    'screen_year': '',
                    'cpu_year': '',
                    'price': ws.cell(r, 8).value or ws.cell(r, 7).value or ws.cell(r, 6).value
                }
            elif any(k in col2_lower for k in ['laptop', 'xách tay', 'macbook', 'dell latitude', 'thinkpad']):
                laptop_in_file += 1
            elif any(k in col2_lower for k in ['máy in', 'máy scan', 'máy quét', 'photocopy', 'canon', 'hp laser']):
                printer_in_file += 1
            else:
                other_in_file += 1
        else:
            # Sub-item (like + Màn hình or + Khối CPU)
            if current_pc:
                col2_lower = col2.lower()
                code_val = col5 if col5 and not col5.isdigit() else col4
                year_val = col3
                if 'màn hình' in col2_lower or 'mh' in col2_lower:
                    current_pc['screen_code'] = code_val
                    current_pc['screen_year'] = str(year_val or '')
                elif 'cpu' in col2_lower or 'khối' in col2_lower or 'thùng' in col2_lower:
                    current_pc['cpu_code'] = code_val
                    current_pc['cpu_year'] = str(year_val or '')

    if current_pc:
        pc_in_file += 1

    print(f"File '{fname}': PC={pc_in_file}, Laptop={laptop_in_file}, Máy in/Scan={printer_in_file}, Khác={other_in_file} | Tổng = {pc_in_file + laptop_in_file + printer_in_file + other_in_file}")
    total_pc_sets += pc_in_file
    total_laptops += laptop_in_file
    total_printers += printer_in_file
    total_others += other_in_file

print("\n==========================================")
print(f"TỔNG CỘNG KHỐI CNTT (BỘ MÁY VI TÍNH HOÀN CHỈNH):")
print(f"  • Bộ máy vi tính (Màn hình + CPU): {total_pc_sets}")
print(f"  • Laptop / Máy tính xách tay: {total_laptops}")
print(f"  • Máy in / Máy Scan: {total_printers}")
print(f"  • Thiết bị mạng / Server / Khác: {total_others}")
print(f"  => TỔNG SỐ THIẾT BỊ CNTT: {total_pc_sets + total_laptops + total_printers + total_others}")
print("==========================================")
