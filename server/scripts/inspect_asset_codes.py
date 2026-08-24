import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def check_file(path):
    print(f"\n==========================================")
    print(f"FILE: {os.path.basename(path)}")
    print(f"==========================================")
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            if ws.max_row < 5:
                continue
            print(f"\n--- Sheet: '{s}' (Rows: {ws.max_row}, Cols: {ws.max_column}) ---")
            
            # Find header rows (first 25 rows)
            for r in range(1, min(ws.max_row + 1, 25)):
                row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
                row_str = [str(v) if v is not None else '' for v in row_vals]
                if any('mã' in v.lower() or 'tài sản' in v.lower() or 'stt' in v.lower() for v in row_str):
                    print(f"  Header row {r}: {row_vals}")
            
            # Print sample data rows
            data_count = 0
            for r in range(1, ws.max_row + 1):
                col1 = ws.cell(r, 1).value
                col2 = ws.cell(r, 2).value
                col3 = ws.cell(r, 3).value
                col4 = ws.cell(r, 4).value
                col5 = ws.cell(r, 5).value
                
                # Check if it looks like a data row
                if col1 is not None and isinstance(col1, (int, float)) or (col1 and str(col1).strip().isdigit()):
                    data_count += 1
                    if data_count <= 8:
                        print(f"  Data row {r} (STT {col1}): Col2='{col2}', Col3='{col3}', Col4='{col4}', Col5='{col5}'")
            print(f"  -> Total data items found: {data_count}")
    except Exception as e:
        print(f"Error reading {path}: {e}")

# 1. Main summary file (Dược & TBYT & Các Khoa)
check_file(r'f:\QLTS\TS\TỔNG HỢP TS CCDC TBYT CDC 2026.xlsx')

# 2. Sample CNTT files
for f in os.listdir(r'f:\QLTS\TS\CNTT')[:5]:
    if f.endswith('.xlsx'):
        check_file(os.path.join(r'f:\QLTS\TS\CNTT', f))

# 3. Sample TBVP files
for f in os.listdir(r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien')[:5]:
    if f.endswith('.xlsx'):
        check_file(os.path.join(r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien', f))
