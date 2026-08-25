import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = r'f:\QLTS\BẢNG THEO DÕI THỜI GIAN SỬA CHỮA, BẢO TRÌ, HIỆU CHUẨN 2026.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

print("ALL SHEET NAMES IN FILE:")
for idx, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f"[{idx+1}] Sheet: '{name}' | MaxRow: {ws.max_row} | MaxCol: {ws.max_column}")
    # print first 10 rows title
    for r in range(1, min(ws.max_row+1, 8)):
        row_v = [ws.cell(r, c).value for c in range(1, min(ws.max_column+1, 15))]
        if any(row_v):
            print(f"    R{r}: {row_v[:8]}")
