import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def analyze_cntt_file(filepath):
    print(f"\n==========================================")
    print(f"CNTT: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for s in wb.sheetnames:
        ws = wb[s]
        if ws.max_row < 5:
            continue
        print(f"--- Sheet: {s} (max_row: {ws.max_row}) ---")
        items = []
        for r in range(1, min(ws.max_row + 1, 35)):
            row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 10))]
            if any(row_vals):
                print(f"  R{r:02d}: {row_vals}")

analyze_cntt_file(r'f:\QLTS\TS\CNTT\1. PKĐK 2026 chốt.xlsx')
analyze_cntt_file(r'f:\QLTS\TS\CNTT\2. XN KK2026 chốt.xlsx')
analyze_cntt_file(r'f:\QLTS\TS\CNTT\14. TCKT 2026 chốt.xlsx')
