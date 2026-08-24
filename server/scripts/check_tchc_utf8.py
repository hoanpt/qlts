import sqlite3
import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

conn = sqlite3.connect(r'f:\QLTS\server\prisma\dev.db')
cursor = conn.cursor()

# 1. Check Department TCHC in DB
cursor.execute("SELECT id, code, name, location FROM Department WHERE code LIKE '%TCHC%' OR name LIKE '%Tổ chức%' OR name LIKE '%Hành chính%'")
depts = cursor.fetchall()
print("Departments matching TCHC in DB:")
for d in depts:
    print(d)
    cursor.execute("SELECT COUNT(*) FROM Asset WHERE departmentId = ?", (d[0],))
    cnt = cursor.fetchone()[0]
    print(f"  -> Asset count in DB for dept {d[0]}: {cnt}")
    
    # Sample assets
    cursor.execute("SELECT id, assetCode, name, categoryId, status, location FROM Asset WHERE departmentId = ? LIMIT 10", (d[0],))
    for a in cursor.fetchall():
        print(f"     Sample: {a}")

conn.close()

# 2. Check files in TS folder relating to TCHC
print("\n--- Checking TCHC Excel files ---")
main_wb = openpyxl.load_workbook(r'f:\QLTS\TS\TỔNG HỢP TS CCDC TBYT CDC 2026.xlsx', data_only=True)
print("Sheets in main workbook:", main_wb.sheetnames)
for s in main_wb.sheetnames:
    if 'TCHC' in s.upper() or 'HC' in s.upper():
        print(f"  Found sheet in main: {s}, rows: {main_wb[s].max_row}, cols: {main_wb[s].max_column}")

print("\nFiles in CNTT matching TCHC:")
for f in os.listdir(r'f:\QLTS\TS\CNTT'):
    if 'TCHC' in f.upper() or 'HC' in f.upper():
        print(f"  CNTT file: {f}")

print("\nFiles in TBVP matching TCHC:")
for f in os.listdir(r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien'):
    if 'TCHC' in f.upper() or 'HC' in f.upper():
        print(f"  TBVP file: {f}")
