import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = r'f:\QLTS\BẢNG THEO DÕI THỜI GIAN SỬA CHỮA, BẢO TRÌ, HIỆU CHUẨN 2026.xlsx'

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    sys.exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
print(f"Workbook sheets: {wb.sheetnames}")

for s_name in wb.sheetnames:
    ws = wb[s_name]
    print(f"\n================ SHEET: '{s_name}' (Rows: {ws.max_row}, Cols: {ws.max_column}) ================")
    for r in range(1, min(ws.max_row + 1, 40)):
        row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 20))]
        if any(v is not None for v in row_vals):
            # Print non-empty rows
            print(f"R{r:2d}: {row_vals}")
