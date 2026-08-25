import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

file_path = r'f:\QLTS\BẢNG THEO DÕI THỜI GIAN SỬA CHỮA, BẢO TRÌ, HIỆU CHUẨN 2026.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)

for sheet_name in ['Sửa chữa', 'Bảo trì', 'Hiệu chuẩn, KĐ, KX']:
    ws = wb[sheet_name]
    print(f"\n=======================================================")
    print(f"SHEET: {sheet_name} (Rows: {ws.max_row}, Cols: {ws.max_column})")
    print(f"=======================================================")
    
    # Header row
    for r in range(1, 10):
        v = str(ws.cell(r, 1).value or '')
        if 'stt' in v.lower():
            headers = [f"Col{c}: {ws.cell(r, c).value}" for c in range(1, ws.max_column+1) if ws.cell(r, c).value]
            print(f"Header at Row {r}:")
            for h in headers:
                print(f"  {h}")
            break

    print("\nSample Data Rows:")
    for r in range(6, min(ws.max_row+1, 25)):
        row_vals = [ws.cell(r, c).value for c in range(1, ws.max_column+1)]
        if any(row_vals):
            print(f"Row {r:2d}: {row_vals[:16]}")
