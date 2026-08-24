import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def print_file_rows(filepath):
    print(f"\n==========================================")
    print(f"FILE: {os.path.basename(filepath)}")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"\n--- Sheet: {s} ---")
        for r in range(1, min(ws.max_row + 1, 35)):
            vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 12))]
            if any(vals):
                print(f"  R{r:02d}: {vals}")

print_file_rows(r'f:\QLTS\TS\CNTT\1. PKDK chốt.xlsx')
print_file_rows(r'f:\QLTS\TS\CNTT\2. XN chốt.xlsx')
print_file_rows(r'f:\QLTS\TS\CNTT\15. TCHC chốt.xlsx')
