import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def inspect_file(filepath):
    print(f"\n=== Inspecting {os.path.basename(filepath)} ===")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"Sheet: {s}, rows: {ws.max_row}, cols: {ws.max_column}")
        if ws.max_row > 5:
            # Print non-empty sample rows
            count = 0
            for r in range(1, min(ws.max_row + 1, 50)):
                row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
                if any(v is not None for v in row_vals):
                    count += 1
                    if count <= 15:
                        print(f"  Row {r}: {row_vals}")

inspect_file(r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien\15 TCHC TS KK T1 2026.xlsx')
inspect_file(r'f:\QLTS\TS\CNTT\15. TCHC chốt.xlsx')
inspect_file(r'f:\QLTS\TS\CNTT\15. TCHC 11.3 - in.xlsx')
