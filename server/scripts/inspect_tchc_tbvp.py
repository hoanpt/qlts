import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien\15 TCHC TS KK T1 2026.xlsx', data_only=True)
for s in wb.sheetnames:
    ws = wb[s]
    print(f"Sheet: {s}, rows: {ws.max_row}, cols: {ws.max_column}")
    items = 0
    for r in range(1, ws.max_row + 1):
        stt = ws.cell(r, 1).value
        name = ws.cell(r, 2).value
        code = ws.cell(r, 3).value
        if stt is not None and name is not None:
            items += 1
            if items <= 15:
                print(f"  Item {stt}: {code} - {name}")
    print(f"Total items in sheet {s}: {items}")
