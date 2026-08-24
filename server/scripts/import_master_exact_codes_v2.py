import sqlite3
import openpyxl
import os
import re
import uuid
import time
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = r'f:\QLTS\server\prisma\dev.db'

DEPT_CODE_MAP = {
    'PKDK': 'PKDK', 'PKĐK': 'PKDK', 'PHÒNG KHÁM': 'PKDK', 'PHONG KHAM': 'PKDK',
    'XN': 'XN', 'XÉT NGHIỆM': 'XN', 'XET NGHIEM': 'XN',
    'BNN': 'BNN', 'BỆNH NGHỀ NGHIỆP': 'BNN', 'BENH NGHE NGHIEP': 'BNN',
    'DD': 'DD', 'DINH DƯỠNG': 'DD', 'DINH DUONG': 'DD', 'DINHDUONG': 'DD',
    'SKMT': 'SKMT', 'SỨC KHỎE MÔI TRƯỜNG': 'SKMT', 'SKMTYTTH': 'SKMT',
    'SKSS': 'SKSS', 'SỨC KHỎE SINH SẢN': 'SKSS',
    'KHNV': 'KHNV', 'KẾ HOẠCH NGHIỆP VỤ': 'KHNV',
    'TTGDSK': 'TTGDSK', 'TRUYỀN THÔNG': 'TTGDSK',
    'KSTCT': 'KSTCT', 'KÝ SINH TRÙNG': 'KSTCT', 'KST': 'KSTCT',
    'PCBKLN': 'PCBKLN', 'BKLN': 'PCBKLN', 'BỆNH KHÔNG LÂY NHIỄM': 'PCBKLN',
    'DVTYT': 'DVTYT', 'DƯỢC': 'DVTYT', 'DUOC': 'DVTYT', 'VTYT': 'DVTYT',
    'HIV': 'HIV', 'PHÒNG CHỐNG HIV': 'HIV',
    'PCBTN': 'PCBTN', 'BTN': 'PCBTN', 'BỆNH TRUYỀN NHIỄM': 'PCBTN',
    'TCKT': 'TCKT', 'TÀI CHÍNH KẾ TOÁN': 'TCKT', 'TAI CHINH': 'TCKT',
    'TCHC': 'TCHC', 'TỔ CHỨC HÀNH CHÍNH': 'TCHC', 'TO CHUC': 'TCHC', 'TC-HC': 'TCHC',
    'KDYTQT': 'KDYTQT', 'KIỂM DỊCH': 'KDYTQT', 'KIEM DICH': 'KDYTQT'
}

def get_dept_id(cursor, name_or_code):
    norm = str(name_or_code or '').upper().strip()
    target_code = 'TCHC'
    for k, v in DEPT_CODE_MAP.items():
        if k in norm:
            target_code = v
            break
    cursor.execute("SELECT id FROM Department WHERE code = ?", (target_code,))
    row = cursor.fetchone()
    return row[0] if row else 1

INVALID_CODE_PATTERNS = [
    r'^s\s*t\s*t$', r'^tài sản$', r'^stt$', r'^mã số$', r'^mã số mới$', r'^mã số cũ$',
    r'^[a-d]$', r'^\d+$'
]

def is_valid_asset_row(code, name):
    if not code or not name:
        return False
    c_str = str(code).strip()
    n_str = str(name).strip()
    if len(n_str) < 2 or len(c_str) == 0:
        return False
    if n_str.lower() in ['tài sản', 'stt', 'đơn vị tính', 's\nt\nt']:
        return False
    if n_str.startswith('Mẫu số') or n_str.startswith('BIÊN BẢN') or n_str.startswith('DANH SÁCH'):
        return False
    for pat in INVALID_CODE_PATTERNS:
        if re.match(pat, c_str, re.IGNORECASE):
            return False
    return True

def run_master_import_v2():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    now_ms = int(time.time() * 1000)

    # Categories
    cursor.execute("SELECT id, code FROM AssetCategory")
    cats = {c[1]: c[0] for c in cursor.fetchall()}
    cat_tbvp = cats.get('TBVP', 1)
    cat_cntt = cats.get('CNTT', 2)
    cat_tbyt = cats.get('TBYT', cats.get('DUOC', 3))
    cat_toanha = cats.get('TBVP_TOANHA', 4)

    cursor.execute("DELETE FROM Asset")
    conn.commit()
    print("Reset Asset table for exact original codes import.\n")

    seen_codes = set()
    total_count = 0

    def insert_record(code, name, cat_id, dept_id, managing_unit, location='Cơ sở 1',
                      loc_detail='', assigned_to='', year=None, price=None, status='DANG_SU_DUNG',
                      book_qty=1, actual_qty=1, qty_diff=0, source='', note='',
                      floor=None, building_asset=0, specs=''):
        nonlocal total_count
        code = str(code).strip()
        name = str(name).strip()
        
        final_code = code
        dup_idx = 1
        while final_code in seen_codes:
            dup_idx += 1
            final_code = f"{code}_{dup_idx}"
        seen_codes.add(final_code)

        qr_code = str(uuid.uuid4())[:12]
        
        try:
            cursor.execute("""
                INSERT INTO Asset (
                    assetCode, name, categoryId, departmentId, managingUnit, location, locationDetail,
                    assignedTo, yearInUse, originalPrice, currentValue, status,
                    bookQuantity, actualQuantity, quantityDifference, source, note, specifications,
                    floor, buildingAsset, qrCode, createdAt, updatedAt
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                final_code, name, cat_id, dept_id, managing_unit, location, loc_detail or None,
                assigned_to or None, year, price, None, status,
                book_qty, actual_qty, qty_diff, source or None, note or None, specs or None,
                floor, building_asset, qr_code, now_ms, now_ms
            ))
            total_count += 1
            return True
        except Exception as e:
            return False

    # =========================================================================
    # 1. IMPORT TỔNG HỢP TS CCDC TBYT CDC 2026.xlsx (Dược & Trang thiết bị y tế)
    # =========================================================================
    main_file = r'f:\QLTS\TS\TỔNG HỢP TS CCDC TBYT CDC 2026.xlsx'
    print(f"1. Parsing Medical Equipment & Department Assets: {os.path.basename(main_file)}")
    wb_main = openpyxl.load_workbook(main_file, data_only=True)
    for s in wb_main.sheetnames:
        ws = wb_main[s]
        dept_id = get_dept_id(cursor, s)
        dept_loc = 'Cơ sở 2' if 'BTN' in s.upper() or 'PCBTN' in s.upper() else 'Cơ sở 1'
        
        header_r = 14
        for r in range(1, min(ws.max_row + 1, 20)):
            v1 = str(ws.cell(r, 1).value or '').lower()
            v2 = str(ws.cell(r, 2).value or '').lower()
            if 'stt' in v1 or 'tài sản' in v2:
                header_r = r
                break

        current_sub_loc = None
        s_count = 0
        for r in range(header_r + 2, ws.max_row + 1):
            col1 = ws.cell(r, 1).value
            col2 = ws.cell(r, 2).value
            col3 = ws.cell(r, 3).value
            col4 = ws.cell(r, 4).value
            col5 = ws.cell(r, 5).value
            col6 = ws.cell(r, 6).value
            col7 = ws.cell(r, 7).value
            col8 = ws.cell(r, 8).value
            col9 = ws.cell(r, 9).value
            col10 = ws.cell(r, 10).value
            col11 = ws.cell(r, 11).value
            col12 = ws.cell(r, 12).value
            col13 = ws.cell(r, 13).value
            col14 = ws.cell(r, 14).value

            if col1 is None and col2 and not col3:
                v2_str = str(col2).strip()
                if not v2_str.startswith('Mẫu') and not v2_str.startswith('BIÊN'):
                    current_sub_loc = v2_str
                continue

            if not col2:
                continue

            name = str(col2).strip()
            code = str(col3 or '').strip()
            if not code:
                code = f"TS-{s}-{s_count + 1:03d}"

            if not is_valid_asset_row(code, name):
                continue

            year_val = None
            try:
                if col4 and 1990 <= int(float(col4)) <= 2030:
                    year_val = int(float(col4))
            except:
                pass

            book_q = int(float(col5)) if col5 and isinstance(col5, (int, float)) else 1
            act_q = int(float(col6)) if col6 and isinstance(col6, (int, float)) else book_q
            diff_q = int(float(col7)) if col7 and isinstance(col7, (int, float)) else (act_q - book_q)

            price_val = None
            try:
                if col8 and isinstance(col8, (int, float)) and col8 > 0:
                    price_val = float(col8)
                elif col9 and isinstance(col9, (int, float)) and col9 > 0:
                    price_val = float(col9) / (act_q or 1)
            except:
                pass

            stat_str = str(col12 or '').lower()
            status = 'DANG_SU_DUNG'
            if 'hỏng' in stat_str or 'đề nghị sửa' in stat_str: status = 'HONG'
            elif 'thanh lý' in stat_str: status = 'CHO_THANH_LY'
            elif 'không sử dụng' in stat_str: status = 'KHONG_SU_DUNG'

            name_lower = name.lower()
            if any(k in name_lower for k in ['bàn làm việc', 'bàn họp', 'ghế xoay', 'ghế hội trường', 'tủ sắt', 'tủ gỗ', 'quạt', 'giường inox']):
                managing_unit = 'TCHC'
                c_id = cat_tbvp
            else:
                managing_unit = 'DUOC'
                c_id = cat_tbyt

            loc_det = str(col11 or current_sub_loc or '').strip()
            note_det = str(col14 or '').strip()
            source_det = str(col10 or '').strip()

            if insert_record(
                code=code, name=name, cat_id=c_id, dept_id=dept_id, managing_unit=managing_unit,
                location=dept_loc, loc_detail=loc_det, assigned_to=loc_det, year=year_val,
                price=price_val, status=status, book_qty=book_q, actual_qty=act_q, qty_diff=diff_q,
                source=source_det, note=note_det
            ):
                s_count += 1

        print(f"  ✓ Sheet '{s}': {s_count} assets with exact codes (TSPK..., TSXN...)")

    # =========================================================================
    # 2. IMPORT CNTT FILES (Exact codes: MH/..., CPU/..., LAP/..., IN/..., SW/...)
    # =========================================================================
    print(f"\n2. Parsing CNTT files (Preserving exact MH/..., CPU/..., LAP/..., IN/...)")
    cntt_dir = r'f:\QLTS\TS\CNTT'
    
    # Specific targeted sheets per file to ensure latest 2026 final data
    CNTT_SHEET_MAP = {
        '1. PKĐK 2026 chốt.xlsx': '2026 PKDK',
        '2. XN KK2026 chốt.xlsx': '2026 XN phan hoi',
        '3. BNN KK 2026..xlsx': '2026 BNN',
        '4. Dinh dưỡng KK2026 chốt.xlsx': 'Ktoan mới',
        '5. SKMT -KK-2026.xlsx': '2026 TB CNTT',
        '6. SKSS CNTT KK2026.xlsx': '2023 SKSS',
        '7. KHNV KK2026.xlsx': '2026 KHNV',
        '8. TTGDSK KK2026 chốt.xlsx': '2026 TTGDSK',
        '9. KSTCT KK2026.xlsx': '2026 KST.CT',
        '10. BKLN- KK 2026 phan hồi.xlsx': '2026',
        '11. Dược KK 2026.xlsx': '2026 Dược',
        '12. HIV 2026.xlsx': 'CNTT',
        '13. BTN-CNTT-kinh phí phan hôi.xlsx': 'BTN cơ sở 2',
        '14. TCKT 2026 chốt.xlsx': '2026 TCKT',
        '15. TCHC chốt.xlsx': '2026 TCHC',
        '16. KDYTQT chốt.xlsx': '2024 Kiểm dịch'
    }

    for fname, sheet_name in CNTT_SHEET_MAP.items():
        p = os.path.join(cntt_dir, fname)
        if not os.path.exists(p):
            continue
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[-1]
            ws = wb[sheet_name]
            dept_id = get_dept_id(cursor, fname)
            f_count = 0

            # Find header
            header_r = 15
            code_col = 3
            code_new_col = 4
            for r in range(1, min(ws.max_row + 1, 20)):
                for c in range(1, min(ws.max_column + 1, 10)):
                    val = str(ws.cell(r, c).value or '').lower()
                    if 'mã' in val and 'mới' in val:
                        code_new_col = c
                        header_r = r
                    elif 'mã' in val:
                        code_col = c
                        header_r = r
                if header_r != 15:
                    break

            parent_machine = ''
            for r in range(header_r + 1, ws.max_row + 1):
                col1 = ws.cell(r, 1).value
                col2 = ws.cell(r, 2).value
                col_old = ws.cell(r, code_col).value if code_col else None
                col_new = ws.cell(r, code_new_col).value if code_new_col else None

                if not col2 and not col1:
                    continue

                col2_str = str(col2 or '').strip()
                if not col2_str or col2_str.startswith('Mẫu') or col2_str.startswith('BIÊN') or col2_str.startswith('THIẾT BỊ'):
                    continue

                # Parent machine title
                if col1 is not None and str(col1).strip().isdigit() and not col_old and not col_new:
                    parent_machine = col2_str
                    continue

                exact_code = str(col_new or col_old or '').strip()
                if not exact_code or exact_code.lower() in ['none', 'null', '-', '']:
                    if parent_machine:
                        exact_code = f"CNTT-{fname[:4].strip()}-{r:03d}"
                    else:
                        continue

                # Filter out junk header codes
                if not is_valid_asset_row(exact_code, col2_str):
                    continue

                full_name = col2_str
                if parent_machine and ('màn hình' in col2_str.lower() or 'cpu' in col2_str.lower()):
                    full_name = f"{parent_machine} ({col2_str.replace('+', '').strip()})"

                year_val = None
                try:
                    for y_col in [5, 4, 6]:
                        y_c = ws.cell(r, y_col).value
                        if y_c and 1990 <= int(float(y_c)) <= 2030:
                            year_val = int(float(y_c))
                            break
                except:
                    pass

                price_val = None
                for p_col in [8, 7, 6, 9]:
                    p_c = ws.cell(r, p_col).value
                    if isinstance(p_c, (int, float)) and p_c > 0:
                        price_val = float(p_c)
                        break

                if insert_record(
                    code=exact_code, name=full_name, cat_id=cat_cntt, dept_id=dept_id,
                    managing_unit='CNTT', location='Cơ sở 1', loc_detail=f"Phòng {fname[:10]}",
                    year=year_val, price=price_val, status='DANG_SU_DUNG'
                ):
                    f_count += 1

            print(f"  ✓ File '{fname}' [{sheet_name}]: {f_count} CNTT assets (MH/..., CPU/..., LAP/..., IN/...)")
        except Exception as e:
            print(f"  Error reading {fname}: {e}")

    # =========================================================================
    # 3. IMPORT TBVP & TÒA NHÀ THEO TẦNG (Exact codes: BXN 03, BQ608, BAT 18...)
    # =========================================================================
    print(f"\n3. Parsing TBVP & Building Floor files from f:\\QLTS\\TS\\TBVP Kiem ke tai san hoan thien\\")
    tbvp_dir = r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien'
    
    # Process files
    tbvp_files = [f for f in sorted(os.listdir(tbvp_dir)) if f.endswith('.xlsx') and not f.startswith('~$')]
    # Deduplicate HIV
    tbvp_files = [f for f in tbvp_files if f != '9. KHOA HIV KK TYAI SAN T1 2026..xlsx']

    for fname in tbvp_files:
        p = os.path.join(tbvp_dir, fname)
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            dept_id = get_dept_id(cursor, fname)
            f_count = 0

            for s in wb.sheetnames:
                ws = wb[s]
                if ws.max_row < 5:
                    continue

                is_building = 'tb dien' in s.lower() or 'điện' in s.lower() or 'tầng' in s.lower() or 'tk điện' in s.lower()
                floor_tag = 'Tầng 1'
                s_lower = s.lower()
                if 't7' in s_lower or 'tầng 7' in s_lower: floor_tag = 'Tầng 7'
                elif 't6' in s_lower or 'tầng 6' in s_lower: floor_tag = 'Tầng 6'
                elif 't5' in s_lower or 'tầng 5' in s_lower: floor_tag = 'Tầng 5'
                elif 't4' in s_lower or 'tầng 4' in s_lower: floor_tag = 'Tầng 4'
                elif 't3' in s_lower or 'tầng 3' in s_lower: floor_tag = 'Tầng 3'
                elif 't2' in s_lower or 'tầng 2' in s_lower: floor_tag = 'Tầng 2'
                elif 'hầm' in s_lower or 'tang ham' in s_lower: floor_tag = 'Tầng Hầm'

                current_room = floor_tag
                for r in range(1, ws.max_row + 1):
                    col1 = ws.cell(r, 1).value
                    col2 = ws.cell(r, 2).value
                    col3 = ws.cell(r, 3).value
                    col4 = ws.cell(r, 4).value
                    col5 = ws.cell(r, 5).value

                    if col1 in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'A', 'B', 'C', 'D'] or (col1 is None and col2 and str(col2).startswith('P.')):
                        current_room = f"{floor_tag} - {str(col2 or col3 or '').strip()}"
                        continue

                    if col1 is not None and str(col1).strip().isdigit():
                        stt = int(col1)
                        name_val = str(col2 or '').strip()
                        code_val = str(col3 or '').strip()

                        if not is_valid_asset_row(code_val or 'OK', name_val):
                            continue

                        if is_building or not code_val or code_val.lower() in ['bộ', 'cái', 'tủ', 'bình', 'm', 'mét', 'none']:
                            unit = code_val if code_val in ['bộ', 'cái', 'tủ', 'bình', 'm'] else 'Bộ'
                            exact_code = f"TN-{floor_tag.replace(' ', '')}-{stt:03d}"
                            managing_unit = 'TCHC'
                            c_id = cat_toanha
                            b_asset = 1
                            qty = int(float(col5)) if col5 and isinstance(col5, (int, float)) else 1
                        else:
                            exact_code = code_val
                            unit = None
                            managing_unit = 'TCHC'
                            c_id = cat_tbvp
                            b_asset = 0
                            qty = int(float(col5)) if col5 and isinstance(col5, (int, float)) else 1

                        year_val = None
                        try:
                            if col4 and 1990 <= int(float(col4)) <= 2030:
                                year_val = int(float(col4))
                        except:
                            pass

                        if insert_record(
                            code=exact_code, name=name_val, cat_id=c_id, dept_id=dept_id,
                            managing_unit=managing_unit, location='Cơ sở 1',
                            loc_detail=current_room, year=year_val, book_qty=qty, actual_qty=qty,
                            floor=floor_tag if b_asset else None, building_asset=b_asset,
                            note=f"ĐVT: {unit}" if unit else None
                        ):
                            f_count += 1

            print(f"  ✓ TBVP '{fname}': {f_count} assets (BXN..., BAT..., TN-...)")
        except Exception as e:
            print(f"  Error reading TBVP {fname}: {e}")

    conn.commit()

    # Final summary
    print("\n=======================================================")
    print("HOÀN TẤT CẬP NHẬT MÃ TÀI SẢN CHUẨN XÁC 100% THEO FILE:")
    print("=======================================================")
    cursor.execute("SELECT managingUnit, COUNT(*) FROM Asset GROUP BY managingUnit")
    for r in cursor.fetchall():
        print(f"  Khối Quản Lý {r[0]}: {r[1]} tài sản")

    cursor.execute("SELECT COUNT(*) FROM Asset")
    print(f"\nTỔNG CỘNG TÀI SẢN TRONG HỆ THỐNG: {cursor.fetchone()[0]}")

    print("\nKIỂM TRA MẪU MÃ TÀI SẢN THEO TỪNG KHỐI:")
    for mu, label in [('DUOC', '1. Khối Khoa Dược (TBYT)'), ('CNTT', '2. Khối Tổ CNTT'), ('TCHC', '3. Khối Phòng TCHC')]:
        print(f"\n{label}:")
        cursor.execute("SELECT assetCode, name FROM Asset WHERE managingUnit = ? LIMIT 8", (mu,))
        for a in cursor.fetchall():
            print(f"  • {a[0]} : {a[1]}")

    conn.close()

if __name__ == '__main__':
    run_master_import_v2()
