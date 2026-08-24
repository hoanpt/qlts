import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

for f in sorted(os.listdir(r'f:\QLTS\TS\CNTT')):
    if f.endswith('.xlsx') and not f.startswith('~$'):
        p = os.path.join(r'f:\QLTS\TS\CNTT', f)
        wb = openpyxl.load_workbook(p, data_only=True)
        sheets_info = [(s, wb[s].max_row) for s in wb.sheetnames]
        print(f"{f}: {sheets_info}")
