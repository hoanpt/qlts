import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

cntt_files = [f for f in os.listdir(r'f:\QLTS\TS\CNTT') if f.endswith('.xlsx') and not f.startswith('~$')]
for fname in sorted(cntt_files)[:5]:
    p = os.path.join(r'f:\QLTS\TS\CNTT', fname)
    wb = openpyxl.load_workbook(p, data_only=True)
    print(f"\n==========================================")
    print(f"CNTT FILE: {fname}")
    for s in wb.sheetnames:
        ws = wb[s]
        print(f"--- Sheet: {s} ---")
        for r in range(1, min(ws.max_row + 1, 25)):
            vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
            if any(vals):
                print(f"  R{r:02d}: {vals}")
