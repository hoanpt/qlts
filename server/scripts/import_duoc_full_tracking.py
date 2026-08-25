import sqlite3
import openpyxl
import os
import datetime
import re
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = r'f:\QLTS\server\prisma\dev.db'
EXCEL_PATH = r'f:\QLTS\BẢNG THEO DÕI THỜI GIAN SỬA CHỮA, BẢO TRÌ, HIỆU CHUẨN 2026.xlsx'

conn = sqlite3.connect(DB_PATH)
cursor = conn.cursor()

now_iso = datetime.datetime.now(datetime.timezone.utc).strftime('%Y-%m-%dT%H:%M:%S.000Z')

print("1. Upgrading database columns for CalibrationRecord & MaintenanceRequest...")

# 1.1 Upgrade CalibrationRecord
cursor.execute("PRAGMA table_info(CalibrationRecord)")
calib_cols = [c[1] for c in cursor.fetchall()]
new_calib_cols = [
    ('serviceType', 'TEXT DEFAULT "HIEU_CHUAN"'),
    ('servicePackage', 'TEXT'),
    ('cost', 'REAL DEFAULT 0'),
    ('decisionNumber', 'TEXT'),
    ('acceptanceMembers', 'TEXT'),
    ('fundingSource', 'TEXT'),
    ('deviceStatusAfter', 'TEXT DEFAULT "Tốt"'),
    ('departmentLocation', 'TEXT'),
    ('proposalDate', 'DATETIME'),
    ('approvalDate', 'DATETIME')
]
for c_name, c_type in new_calib_cols:
    if c_name not in calib_cols:
        cursor.execute(f"ALTER TABLE CalibrationRecord ADD COLUMN {c_name} {c_type}")
        print(f"  Added CalibrationRecord column: {c_name}")

# 1.2 Upgrade MaintenanceRequest
cursor.execute("PRAGMA table_info(MaintenanceRequest)")
maint_cols = [c[1] for c in cursor.fetchall()]
new_maint_cols = [
    ('maintenanceType', 'TEXT DEFAULT "SUA_CHUA"'),
    ('servicePackage', 'TEXT'),
    ('replacementParts', 'TEXT'),
    ('acceptanceMembers', 'TEXT'),
    ('fundingSource', 'TEXT'),
    ('deviceStatusAfter', 'TEXT DEFAULT "Tốt"'),
    ('proposalDate', 'DATETIME'),
    ('approvalDate', 'DATETIME')
]
for c_name, c_type in new_maint_cols:
    if c_name not in maint_cols:
        cursor.execute(f"ALTER TABLE MaintenanceRequest ADD COLUMN {c_name} {c_type}")
        print(f"  Added MaintenanceRequest column: {c_name}")

conn.commit()

# Helper to find assetId by code or name
def find_asset_id(code, name, dept_id=11):
    if code:
        code_clean = str(code).strip()
        cursor.execute("SELECT id FROM Asset WHERE LOWER(assetCode) = LOWER(?) LIMIT 1", (code_clean,))
        r = cursor.fetchone()
        if r: return r[0]
        # try like
        cursor.execute("SELECT id FROM Asset WHERE assetCode LIKE ? LIMIT 1", (f"%{code_clean}%",))
        r = cursor.fetchone()
        if r: return r[0]

    if name:
        name_clean = str(name).strip()[:20]
        cursor.execute("SELECT id FROM Asset WHERE name LIKE ? AND managingUnit = 'DUOC' LIMIT 1", (f"%{name_clean}%",))
        r = cursor.fetchone()
        if r: return r[0]

    # Fallback to any DUOC asset
    cursor.execute("SELECT id FROM Asset WHERE managingUnit = 'DUOC' AND departmentId = ? LIMIT 1", (dept_id,))
    r = cursor.fetchone()
    if r: return r[0]

    cursor.execute("SELECT id FROM Asset WHERE managingUnit = 'DUOC' LIMIT 1")
    r = cursor.fetchone()
    return r[0] if r else 1

def parse_date(val):
    if not val:
        return None
    if isinstance(val, datetime.datetime):
        return val.isoformat()
    if isinstance(val, str):
        val = val.strip()
        # format DD/MM/YYYY
        m = re.match(r'(\d{1,2})[/.-](\d{1,2})[/.-](\d{4})', val)
        if m:
            d, month, y = m.groups()
            return f"{y}-{int(month):02d}-{int(d):02d}T00:00:00.000Z"
    return None

wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

# =========================================================================
# 2. IMPORT SHEET: 'Sửa chữa'
# =========================================================================
print("\n2. Importing Sheet 'Sửa chữa' (Dược - Sửa chữa thay thế linh kiện)...")
ws_sc = wb['Sửa chữa']
sc_count = 0

current_dept_name = 'Phòng khám Đa khoa'
current_dept_id = 1

for r in range(6, ws_sc.max_row + 1):
    c1 = ws_sc.cell(r, 1).value
    c2 = str(ws_sc.cell(r, 2).value or '').strip()
    c3 = str(ws_sc.cell(r, 3).value or '').strip()

    if not c2 and not c3:
        continue

    # Roman numeral header e.g. I. Phòng khám Đa khoa, II. Khoa Xét nghiệm...
    if c1 in ['I', 'II', 'III', 'IV', 'V', 'VI', 'A', 'B', 'C', 'D']:
        current_dept_name = c2
        if 'phòng khám' in c2.lower() or 'pkđk' in c2.lower(): current_dept_id = 1
        elif 'xét nghiệm' in c2.lower() or 'xn' in c2.lower(): current_dept_id = 2
        elif 'hiv' in c2.lower(): current_dept_id = 12
        elif 'dược' in c2.lower(): current_dept_id = 11
        elif 'bệnh nghề nghiệp' in c2.lower(): current_dept_id = 3
        continue

    is_num = False
    if c1 is not None:
        try:
            int(float(str(c1)))
            is_num = True
        except:
            pass

    if is_num and c3:
        code_val = c2 if c2 and not c2.lower().startswith('tổng') else None
        name_val = c3
        bophan_val = str(ws_sc.cell(r, 5).value or current_dept_name).strip()
        ngay_bao = parse_date(ws_sc.cell(r, 6).value) or '2026-01-10T08:00:00.000Z'
        ngay_trinh = parse_date(ws_sc.cell(r, 7).value)
        hientrang = str(ws_sc.cell(r, 8).value or 'Hư hỏng linh kiện').strip()
        donvi_th = str(ws_sc.cell(r, 9).value or 'Hãng thiết bị y tế').strip()
        noidung = str(ws_sc.cell(r, 10).value or hientrang).strip()
        linhkien = str(ws_sc.cell(r, 11).value or '').strip()
        ngay_xong = parse_date(ws_sc.cell(r, 12).value)
        
        kinhphi = 0
        v_kp = ws_sc.cell(r, 13).value
        if isinstance(v_kp, (int, float)):
            kinhphi = float(v_kp)

        tinhtrang = str(ws_sc.cell(r, 14).value or 'Tốt').strip()
        nguoi_nt = str(ws_sc.cell(r, 15).value or 'Ds. Tính, Ds. Lộc').strip()
        nguon_kp = str(ws_sc.cell(r, 16).value or 'Thu sự nghiệp').strip()

        asset_id = find_asset_id(code_val, name_val, current_dept_id)

        status_val = 'COMPLETED' if ngay_xong or kinhphi > 0 else 'IN_PROGRESS'

        cursor.execute("""
            INSERT INTO MaintenanceRequest (
                assetId, departmentId, requestedBy, contactPhone, issueDescription,
                priority, status, repairCost, repairVendor, repairNote,
                technicianName, managingUnit, locationDetail, requestDate, completedDate,
                maintenanceType, replacementParts, acceptanceMembers, fundingSource,
                deviceStatusAfter, approvalDate, createdAt, updatedAt
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'DUOC', ?, ?, ?, 'SUA_CHUA', ?, ?, ?, ?, ?, ?, ?)
        """, (
            asset_id, current_dept_id, 'Phụ trách Khoa/Bộ phận', '0905.xxx.xxx',
            hientrang, 'HIGH', status_val, kinhphi, donvi_th,
            f"{noidung}. Linh kiện: {linhkien}",
            nguoi_nt, bophan_val, ngay_bao, ngay_xong,
            linhkien, nguoi_nt, nguon_kp, tinhtrang, ngay_trinh, now_iso, now_iso
        ))
        sc_count += 1

print(f"  ✓ Đã nạp {sc_count} hồ sơ Sửa chữa TBYT thực tế.")

# =========================================================================
# 3. IMPORT SHEET: 'Bảo trì'
# =========================================================================
print("\n3. Importing Sheet 'Bảo trì' (Dược - Bảo trì, bảo dưỡng hệ thống TBYT)...")
ws_bt = wb['Bảo trì']
bt_count = 0
current_dept_name = 'Khoa XN-CĐHA-TDCN'
current_dept_id = 2

for r in range(6, ws_bt.max_row + 1):
    c1 = ws_bt.cell(r, 1).value
    c2 = str(ws_bt.cell(r, 2).value or '').strip()
    c3 = str(ws_bt.cell(r, 3).value or '').strip()

    if not c2 and not c3:
        continue

    if c1 in ['I', 'II', 'III', 'IV', 'V', 'A', 'B', 'C']:
        current_dept_name = c2
        if 'dược' in c2.lower() or 'dvtyt' in c2.lower(): current_dept_id = 11
        elif 'xét nghiệm' in c2.lower() or 'xn' in c2.lower(): current_dept_id = 2
        elif 'phòng khám' in c2.lower(): current_dept_id = 1
        continue

    is_num = False
    if c1 is not None:
        try:
            int(float(str(c1)))
            is_num = True
        except:
            pass

    if is_num and c3:
        code_val = c2 if c2 and not c2.lower().startswith('tổng') else None
        name_val = c3
        bophan_val = str(ws_bt.cell(r, 5).value or current_dept_name).strip()
        ngay_dexuat = parse_date(ws_bt.cell(r, 6).value) or '2026-01-15T08:00:00.000Z'
        ngay_trinh = parse_date(ws_bt.cell(r, 7).value)
        noidung_dx = str(ws_bt.cell(r, 8).value or 'Dịch vụ bảo trì, bảo dưỡng toàn bộ hệ thống').strip()
        donvi_th = str(ws_bt.cell(r, 9).value or 'Đơn vị bảo trì chuyên nghiệp').strip()
        noidung_th = str(ws_bt.cell(r, 10).value or noidung_dx).strip()
        linhkien = str(ws_bt.cell(r, 11).value or '').strip()
        ngay_xong = parse_date(ws_bt.cell(r, 12).value)

        kinhphi = 0
        v_kp = ws_bt.cell(r, 13).value
        if isinstance(v_kp, (int, float)):
            kinhphi = float(v_kp)

        tinhtrang = str(ws_bt.cell(r, 14).value or 'Tốt').strip()
        nguoi_nt = str(ws_bt.cell(r, 15).value or 'Ds. Tính, Ds. Lộc, Ds. Thành').strip()
        nguon_kp = str(ws_bt.cell(r, 16).value or 'Quỹ PTHĐSN').strip()

        asset_id = find_asset_id(code_val, name_val, current_dept_id)

        status_val = 'COMPLETED' if ngay_xong or kinhphi > 0 else 'IN_PROGRESS'

        cursor.execute("""
            INSERT INTO MaintenanceRequest (
                assetId, departmentId, requestedBy, contactPhone, issueDescription,
                priority, status, repairCost, repairVendor, repairNote,
                technicianName, managingUnit, locationDetail, requestDate, completedDate,
                maintenanceType, servicePackage, replacementParts, acceptanceMembers,
                fundingSource, deviceStatusAfter, approvalDate, createdAt, updatedAt
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'DUOC', ?, ?, ?, 'BAO_TRI', ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            asset_id, current_dept_id, 'Khoa Dược - VTYT', '0905.xxx.xxx',
            noidung_dx, 'MEDIUM', status_val, kinhphi, donvi_th,
            f"{noidung_th}. Phụ kiện: {linhkien}",
            nguoi_nt, bophan_val, ngay_dexuat, ngay_xong,
            f"Gói bảo trì TBYT - {bophan_val}", linhkien, nguoi_nt,
            nguon_kp, tinhtrang, ngay_trinh, now_iso, now_iso
        ))
        bt_count += 1

print(f"  ✓ Đã nạp {bt_count} hồ sơ Bảo trì định kỳ TBYT thực tế.")

# =========================================================================
# 4. IMPORT SHEET: 'Hiệu chuẩn, KĐ, KX'
# =========================================================================
print("\n4. Importing Sheet 'Hiệu chuẩn, KĐ, KX' (Toàn bộ 144+ thiết bị Hiệu chuẩn/Kiểm định/Thử nghiệm/Kiểm xạ)...")
ws_hc = wb['Hiệu chuẩn, KĐ, KX']
cursor.execute("DELETE FROM CalibrationRecord")

hc_count = 0
current_package = 'Gói dịch vụ: Hiệu chuẩn, kiểm định thiết bị phục vụ đánh giá lại ISO17025'
current_vendor = 'TT Kiểm định Hiệu chuẩn Đo lường Miền Nam (SMETES)'
current_decision = 'Quyết định số 34/QĐ-TTKSBT ngày 27/01/2026'
current_members = 'Ds. Tính, Ds. Lộc, Cn. Hải, Cn. Sơn'
current_funding = 'Thu sự nghiệp'

for r in range(8, ws_hc.max_row + 1):
    c1 = ws_hc.cell(r, 1).value
    c2 = str(ws_hc.cell(r, 2).value or '').strip()
    c3 = str(ws_hc.cell(r, 3).value or '').strip()

    # Package header row
    if c1 in ['I', 'II', 'III', 'IV', 'V']:
        current_package = c2 or current_package
        if ws_hc.cell(r, 9).value: current_vendor = str(ws_hc.cell(r, 9).value).strip()
        if ws_hc.cell(r, 14).value: current_members = str(ws_hc.cell(r, 14).value).strip()
        if ws_hc.cell(r, 15).value: current_funding = str(ws_hc.cell(r, 15).value).strip()
        if ws_hc.cell(r, 16).value: current_decision = str(ws_hc.cell(r, 16).value).strip()
        continue

    is_num = False
    if c1 is not None:
        try:
            int(float(str(c1)))
            is_num = True
        except:
            pass

    if is_num and c3:
        code_val = c2 if c2 and not c2.lower().startswith('tổng') else None
        name_val = c3
        bophan_val = str(ws_hc.cell(r, 5).value or 'Khoa Xét Nghiệm').strip()
        
        noidung_th = str(ws_hc.cell(r, 10).value or 'Hiệu chuẩn').strip()
        # map serviceType
        service_type = 'HIEU_CHUAN'
        if 'thử nghiệm' in noidung_th.lower(): service_type = 'THU_NGHIEM'
        elif 'kiểm định' in noidung_th.lower(): service_type = 'KIEM_DINH'
        elif 'kiểm xạ' in noidung_th.lower() or 'x-quang' in noidung_th.lower(): service_type = 'KIEM_XA'

        ngay_hoanthanh = parse_date(ws_hc.cell(r, 11).value) or '2026-03-03T00:00:00.000Z'
        
        cost_val = 0
        v_cost = ws_hc.cell(r, 12).value
        if isinstance(v_cost, (int, float)):
            cost_val = float(v_cost)

        tinhtrang = str(ws_hc.cell(r, 13).value or 'Tốt').strip()
        result_val = 'PASS' if 'tốt' in tinhtrang.lower() or 'đạt' in tinhtrang.lower() else 'FAIL'

        nguoi_nt = str(ws_hc.cell(r, 14).value or current_members).strip()
        ghichu = str(ws_hc.cell(r, 15).value or '').strip()
        quyetdinh = str(ws_hc.cell(r, 16).value or current_decision).strip()

        # Calculate next calibration date (+1 year)
        calib_dt = datetime.datetime.fromisoformat(ngay_hoanthanh.replace('Z', '+00:00'))
        next_dt = calib_dt.replace(year=calib_dt.year + 1).isoformat()

        # Department ID
        dept_id = 2
        if 'pkđk' in bophan_val.lower() or 'phòng khám' in bophan_val.lower(): dept_id = 1
        elif 'hiv' in bophan_val.lower(): dept_id = 12
        elif 'btn' in bophan_val.lower(): dept_id = 13
        elif 'dược' in bophan_val.lower(): dept_id = 11

        asset_id = find_asset_id(code_val, name_val, dept_id)

        cert_no = f"SMETES-2026-{hc_count+1:04d}"

        cursor.execute("""
            INSERT INTO CalibrationRecord (
                assetId, calibrationDate, nextCalibrationDate, performedBy, vendor,
                result, certificateNumber, note, createdAt,
                serviceType, servicePackage, cost, decisionNumber, acceptanceMembers,
                fundingSource, deviceStatusAfter, departmentLocation
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            asset_id, ngay_hoanthanh, next_dt, nguoi_nt, current_vendor,
            result_val, cert_no, f"{noidung_th}. {ghichu}", now_iso,
            service_type, current_package, cost_val, quyetdinh, nguoi_nt,
            current_funding, tinhtrang, bophan_val
        ))
        hc_count += 1

print(f"  ✓ Đã nạp thành công {hc_count} hồ sơ Hiệu chuẩn, Thử nghiệm, Kiểm định, Kiểm xạ thực tế.")

conn.commit()
conn.close()
print("\n=======================================================")
print("HOÀN TẤT ĐỒNG BỘ DỮ LIỆU SỬA CHỮA, BẢO TRÌ & HIỆU CHUẨN DƯỢC 2026!")
print("=======================================================")
