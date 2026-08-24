import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

print("=== INSPECTING CNTT FILES ASSET CODES ===")
cntt_dir = r'f:\QLTS\TS\CNTT'
for f in sorted(os.listdir(cntt_dir)):
    if f.endswith('.xlsx') and not f.startswith('~$'):
        p = os.path.join(cntt_dir, f)
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            print(f"\nFile: {f}")
            for s in wb.sheetnames:
                ws = wb[s]
                if ws.max_row < 5:
                    continue
                # find header
                code_col = None
                code_new_col = None
                name_col = None
                header_r = None
                for r in range(1, min(ws.max_row + 1, 20)):
                    row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
                    for c_idx, val in enumerate(row_vals, 1):
                        val_str = str(val or '').lower()
                        if 'tên' in val_str or 'tài sản' in val_str:
                            name_col = c_idx
                        if 'mã' in val_str and 'mới' in val_str:
                            code_new_col = c_idx
                        elif 'mã' in val_str:
                            code_col = c_idx
                    if name_col and (code_col or code_new_col):
                        header_r = r
                        break
                print(f"  Sheet: '{s}' (Rows: {ws.max_row}), Header at {header_r}: NameCol={name_col}, CodeCol={code_col}, CodeNewCol={code_new_col}")
                # print 3 sample items
                samples = []
                for r in range((header_r or 15) + 1, min(ws.max_row + 1, (header_r or 15) + 20)):
                    stt = ws.cell(r, 1).value
                    name = ws.cell(r, name_col or 2).value if name_col else ws.cell(r, 2).value
                    c_old = ws.cell(r, code_col).value if code_col else None
                    c_new = ws.cell(r, code_new_col).value if code_new_col else None
                    if name and stt is not None:
                        samples.append(f"STT {stt}: name='{str(name)[:30]}', code='{c_old}', code_new='{c_new}'")
                        if len(samples) >= 3:
                            break
                for sp in samples:
                    print(f"    {sp}")
        except Exception as e:
            print(f"  Error: {e}")

print("\n=== INSPECTING TBVP FILES ASSET CODES ===")
tbvp_dir = r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien'
for f in sorted(os.listdir(tbvp_dir)):
    if f.endswith('.xlsx') and not f.startswith('~$'):
        p = os.path.join(tbvp_dir, f)
        try:
            wb = openpyxl.load_workbook(p, data_only=True)
            print(f"\nFile: {f}")
            for s in wb.sheetnames:
                ws = wb[s]
                if ws.max_row < 5:
                    continue
                # find header
                code_col = None
                name_col = None
                header_r = None
                for r in range(1, min(ws.max_row + 1, 20)):
                    row_vals = [ws.cell(r, c).value for c in range(1, min(ws.max_column + 1, 15))]
                    for c_idx, val in enumerate(row_vals, 1):
                        val_str = str(val or '').lower()
                        if 'tên' in val_str or 'tài sản' in val_str:
                            name_col = c_idx
                        elif 'mã' in val_str:
                            code_col = c_idx
                    if name_col and code_col:
                        header_r = r
                        break
                print(f"  Sheet: '{s}' (Rows: {ws.max_row}): NameCol={name_col}, CodeCol={code_col}")
                samples = []
                for r in range((header_r or 15) + 1, min(ws.max_row + 1, (header_r or 15) + 20)):
                    stt = ws.cell(r, 1).value
                    name = ws.cell(r, name_col or 2).value if name_col else ws.cell(r, 2).value
                    c_val = ws.cell(r, code_col).value if code_col else None
                    if name and stt is not None:
                        samples.append(f"STT {stt}: name='{str(name)[:30]}', code='{c_val}'")
                        if len(samples) >= 3:
                            break
                for sp in samples:
                    print(f"    {sp}")
        except Exception as e:
            print(f"  Error: {e}")
