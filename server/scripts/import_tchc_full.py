import sqlite3
import openpyxl
import os
import re
import uuid
import time
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

DB_PATH = r'f:\QLTS\server\prisma\dev.db'

def import_tchc_all():
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()
    
    # Get TCHC department ID
    cursor.execute("SELECT id FROM Department WHERE code = 'TCHC'")
    row = cursor.fetchone()
    if not row:
        print("TCHC department not found!")
        return
    tchc_id = row[0]

    # Category IDs
    cursor.execute("SELECT id FROM AssetCategory WHERE code = 'TBVP'")
    tbvp_cat_id = cursor.fetchone()[0]
    cursor.execute("SELECT id FROM AssetCategory WHERE code = 'CNTT'")
    cntt_cat_id = cursor.fetchone()[0]

    now_ms = int(time.time() * 1000)
    imported = 0
    updated = 0

    print("=== IMPORTING ALL TCHC ASSETS ===")

    # 1. Import from 15 TCHC TS KK T1 2026.xlsx (All Floor Equipment Sheets)
    tbvp_file = r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien\15 TCHC TS KK T1 2026.xlsx'
    if os.path.exists(tbvp_file):
        wb = openpyxl.load_workbook(tbvp_file, data_only=True)
        for s in wb.sheetnames:
            ws = wb[s]
            if ws.max_row <= 2:
                continue
            
            floor_name = s.replace('TB DIEN 2022', '').strip()
            if not floor_name:
                floor_name = s
            
            current_room = floor_name
            for r in range(1, ws.max_row + 1):
                col1 = ws.cell(r, 1).value
                col2 = ws.cell(r, 2).value
                col3 = ws.cell(r, 3).value
                col4 = ws.cell(r, 4).value

                # If it's a section/room header (e.g. "I: Phòng kỹ thuật X6-9/Y4")
                if col1 in ['I', 'II', 'III', 'IV', 'V', 'VI', 'VII', 'VIII', 'IX', 'X', 'A', 'B', 'C', 'D'] or (col1 is None and col2 and str(col2).startswith('P.')):
                    current_room = f"{floor_name} - {str(col2 or col3 or '').strip()}"
                    continue

                # Check if it's a valid equipment row
                if col1 is not None and str(col1).strip().isdigit():
                    stt = int(col1)
                    unit = str(col2 or '').strip() if col2 else None
                    name = str(col3 or '').strip() if col3 else None
                    
                    # Sometimes col2 is name, col3 is unit
                    if not name and unit:
                        name = unit
                        unit = None
                    elif unit and name and len(unit) > len(name):
                        # swap if unit is longer (likely description)
                        name, unit = unit, name

                    if not name or len(name) < 2:
                        continue

                    # Asset code
                    asset_code = f"TCHC-{floor_name.replace(' ', '')}-{stt:03d}"
                    
                    # Check if already exists by assetCode
                    cursor.execute("SELECT id FROM Asset WHERE assetCode = ?", (asset_code,))
                    existing = cursor.fetchone()
                    if existing:
                        continue

                    qr_code = str(uuid.uuid4())[:12]
                    
                    # Quantity
                    qty = 1
                    for q_col in [4, 5, 6]:
                        val = ws.cell(r, q_col).value
                        if isinstance(val, (int, float)) and val > 0:
                            qty = int(val)
                            break

                    try:
                        cursor.execute("""
                            INSERT INTO Asset (
                                assetCode, name, categoryId, departmentId, location, locationDetail,
                                assignedTo, yearInUse, originalPrice, currentValue, status,
                                bookQuantity, actualQuantity, quantityDifference, source, note, qrCode,
                                createdAt, updatedAt
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            asset_code, name, tbvp_cat_id, tchc_id, 'Cơ sở 1',
                            current_room, 'Phòng TCHC', 2022, None, None, 'DANG_SU_DUNG',
                            qty, qty, 0, 'DA ĐT Tòa nhà CDC', f"ĐVT: {unit or 'Bộ'}", qr_code,
                            now_ms, now_ms
                        ))
                        imported += 1
                    except Exception as e:
                        pass

    # 2. Import from 15. TCHC chốt.xlsx (CNTT Equipment)
    cntt_file = r'f:\QLTS\TS\CNTT\15. TCHC chốt.xlsx'
    if os.path.exists(cntt_file):
        wb = openpyxl.load_workbook(cntt_file, data_only=True)
        for s in ['2026 TCHC', '2024 TCHC', wb.sheetnames[0]]:
            if s in wb.sheetnames:
                ws = wb[s]
                for r in range(16, ws.max_row + 1):
                    stt = ws.cell(r, 1).value
                    name = ws.cell(r, 2).value
                    code_old = ws.cell(r, 3).value
                    code_new = ws.cell(r, 4).value
                    year = ws.cell(r, 5).value

                    if not name or stt is None:
                        continue

                    name = str(name).strip()
                    asset_code = str(code_new or code_old or '').strip()
                    if not asset_code:
                        asset_code = f"CNTT-TCHC-{stt:03d}"

                    cursor.execute("SELECT id FROM Asset WHERE assetCode = ?", (asset_code,))
                    if cursor.fetchone():
                        continue

                    year_in_use = None
                    try:
                        if year and 1990 <= int(year) <= 2030:
                            year_in_use = int(year)
                    except:
                        pass

                    price = None
                    for p_col in [8, 7, 6]:
                        p_val = ws.cell(r, p_col).value
                        if isinstance(p_val, (int, float)) and p_val > 0:
                            price = float(p_val)
                            break

                    qr_code = str(uuid.uuid4())[:12]
                    try:
                        cursor.execute("""
                            INSERT INTO Asset (
                                assetCode, name, categoryId, departmentId, location, locationDetail,
                                assignedTo, yearInUse, originalPrice, currentValue, status,
                                bookQuantity, actualQuantity, quantityDifference, source, note, qrCode,
                                createdAt, updatedAt
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (
                            asset_code, name, cntt_cat_id, tchc_id, 'Cơ sở 1',
                            'Phòng TCHC', 'Phòng TCHC', year_in_use, price, None, 'DANG_SU_DUNG',
                            1, 1, 0, 'Ngân sách / DA CDC', None, qr_code,
                            now_ms, now_ms
                        ))
                        imported += 1
                    except Exception as e:
                        pass
                break

    conn.commit()

    # Total assets in TCHC now
    cursor.execute("SELECT COUNT(*) FROM Asset WHERE departmentId = ?", (tchc_id,))
    total_tchc = cursor.fetchone()[0]

    cursor.execute("SELECT COUNT(*) FROM Asset")
    total_all = cursor.fetchone()[0]

    conn.close()

    print(f"\nImported {imported} additional TCHC assets!")
    print(f"Total TCHC assets in DB: {total_tchc}")
    print(f"Total assets across all departments: {total_all}")

if __name__ == '__main__':
    import_tchc_all()
