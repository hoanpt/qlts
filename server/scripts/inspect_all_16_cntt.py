import openpyxl
import os
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

cntt_dir = r'f:\QLTS\TS\CNTT'
files = [f for f in sorted(os.listdir(cntt_dir)) if f.endswith('.xlsx') and not f.startswith('~$')]

print("=== INSPECTING ALL 16 CNTT FILES ===")
for fname in files:
    p = os.path.join(cntt_dir, fname)
    wb = openpyxl.load_workbook(p, data_only=True)
    print(f"\nFile: {fname} | Sheets: {wb.sheetnames}")
    # inspect chosen sheet
    sheet_name = wb.sheetnames[0]
    for s in wb.sheetnames:
        if any(k in s.lower() for k in ['2026', 'chốt', 'cntt', '2025', '2024', 'ktoan', 'btn']):
            sheet_name = s
            break
    ws = wb[sheet_name]
    print(f"  Chosen sheet: '{sheet_name}' (Rows: {ws.max_row})")
    for r in range(12, min(ws.max_row + 1, 22)):
        c1 = ws.cell(r, 1).value
        c2 = ws.cell(r, 2).value
        c3 = ws.cell(r, 3).value
        c4 = ws.cell(r, 4).value
        c5 = ws.cell(r, 5).value
        if c1 or c2 or c3 or c4:
            print(f"    R{r:2d}: [{c1}] [{c2}] [{c3}] [{c4}] [{c5}]")
