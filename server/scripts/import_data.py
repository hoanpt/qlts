"""
Script import dữ liệu từ Excel vào database SQLite cho hệ thống QLTS CDC Đà Nẵng.
Chạy sau khi backend đã setup xong database.

Usage: python f:\QLTS\server\scripts\import_data.py
"""
import sqlite3
import openpyxl
import os
import re
import uuid
from datetime import datetime

DB_PATH = r'f:\QLTS\server\prisma\dev.db'

# Map department names/abbreviations to their codes
DEPT_MAP = {
    'PKĐK': 'PKDK', 'P. Tư vấn': 'PKDK', 'Phòng khám đa khoa': 'PKDK', 'PKDK': 'PKDK',
    'Khoa XN-CĐHA-TDCN': 'XN', 'XN': 'XN', 'Hóa TP': 'XN', 'Hóa Lý': 'XN',
    'Khoa BNN': 'BNN', 'BNN': 'BNN',
    'Khoa DD': 'DD', 'DD': 'DD', 'Dinh dưỡng': 'DD',
    'SKMT': 'SKMT', 'Khoa SKMT': 'SKMT',
    'Khoa SKSS': 'SKSS', 'SKSS': 'SKSS',
    'KHNV': 'KHNV', 'Phòng KHNV': 'KHNV',
    'TTGDSK': 'TTGDSK', 'Khoa TTGDSK': 'TTGDSK',
    'KSTCT': 'KSTCT', 'K. KST-CT': 'KSTCT', 'KST': 'KSTCT',
    'PCBKLN': 'PCBKLN', 'Khoa PCBKLN': 'PCBKLN', 'KLN': 'PCBKLN',
    'DVTYT': 'DVTYT', 'Dược': 'DVTYT', 'Khoa Dược': 'DVTYT',
    'HIV': 'HIV', 'Khoa HIV': 'HIV', 'Methadone 1': 'HIV', 'Methadone 2': 'HIV',
    'PCBTN': 'PCBTN', 'Khoa PCBTN': 'PCBTN', 'BTN': 'PCBTN', 'PCBTN-TSN': 'PCBTN',
    'TCKT': 'TCKT', 'Phòng TCKT': 'TCKT', 'TC-KT': 'TCKT',
    'TCHC': 'TCHC', 'TC-HC': 'TCHC', 'Phòng TCHC': 'TCHC',
    'KDYTQT': 'KDYTQT', 'KDTYQT': 'KDYTQT',
}

# Status mapping
STATUS_MAP = {
    'Đang sử dụng': 'DANG_SU_DUNG',
    'Đang sử dụng ': 'DANG_SU_DUNG',
    'Hỏng': 'HONG',
    'Hỏng, ĐN sửa': 'HONG',
    'Không sử dụng': 'KHONG_SU_DUNG',
    'ĐN thanh lý': 'CHO_THANH_LY',
    'Đề nghị thanh lý': 'CHO_THANH_LY',
    'Chờ thanh lý': 'CHO_THANH_LY',
    'Đã thanh lý': 'DA_THANH_LY',
    'Bảo trì': 'BAO_TRI',
    'Đang sửa chữa': 'BAO_TRI',
}

def get_status(text):
    if not text:
        return 'DANG_SU_DUNG'
    text = str(text).strip()
    for key, val in STATUS_MAP.items():
        if key.lower() in text.lower():
            return val
    if 'hỏng' in text.lower():
        return 'HONG'
    if 'thanh lý' in text.lower():
        return 'CHO_THANH_LY'
    if 'sử dụng' in text.lower():
        return 'DANG_SU_DUNG'
    return 'DANG_SU_DUNG'

def clean_text(val):
    if val is None:
        return None
    text = str(val).strip()
    text = text.replace('\n', ' ').replace('\r', '')
    return text if text else None

def safe_int(val):
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None

def safe_float(val):
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def find_dept_id(conn, dept_text):
    """Find department ID from department text"""
    if not dept_text:
        return None
    dept_text = str(dept_text).strip()
    
    # Try direct mapping
    for key, code in DEPT_MAP.items():
        if key.lower() in dept_text.lower():
            cursor = conn.execute("SELECT id FROM Department WHERE code = ?", (code,))
            row = cursor.fetchone()
            if row:
                return row[0]
    
    # Try matching by code
    cursor = conn.execute("SELECT id FROM Department WHERE code = ? OR name LIKE ?", 
                          (dept_text, f'%{dept_text}%'))
    row = cursor.fetchone()
    return row[0] if row else None

def get_category_id(conn, category_code):
    cursor = conn.execute("SELECT id FROM AssetCategory WHERE code = ?", (category_code,))
    row = cursor.fetchone()
    return row[0] if row else 1

def import_tbvp_main(conn):
    """Import tài sản TBVP từ file tổng hợp"""
    filepath = r'f:\QLTS\TS\TỔNG HỢP TS CCDC TBYT CDC 2026.xlsx'
    if not os.path.exists(filepath):
        print(f"File not found: {filepath}")
        return
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    category_id = get_category_id(conn, 'TBVP')
    imported = 0
    skipped = 0
    
    # Sheet to department mapping
    sheet_dept_map = {
        'PKĐK': 'PKDK', 'XN': 'XN', 'KDYTQT': 'KDYTQT', 'DVTYT': 'DVTYT',
        'HIV': 'HIV', 'PCBTN': 'PCBTN', 'SKSS': 'SKSS', 'KSTCT': 'KSTCT',
        'BNN': 'BNN', 'DD': 'DD', 'SKMT': 'SKMT', 'TTGDSK': 'TTGDSK',
        'PCBKLN': 'PCBKLN', 'TCKT': 'TCKT', 'TCHC': 'TCHC', 'KHNV': 'KHNV',
    }
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        dept_code = sheet_dept_map.get(sheet_name)
        if not dept_code:
            continue
        
        cursor = conn.execute("SELECT id FROM Department WHERE code = ?", (dept_code,))
        dept_row = cursor.fetchone()
        if not dept_row:
            print(f"  Department not found: {dept_code}")
            continue
        dept_id = dept_row[0]
        
        # Find header row (look for 'STT' or 'A' column patterns)
        header_row = None
        data_start_row = None
        col_map = {}
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), 1):
            vals = [cell.value for cell in row]
            # Look for the data column labels row (A, B, C, D, 1, 2, 3...)
            if vals and vals[0] == 'A' and vals[1] == 'B':
                data_start_row = row_idx + 1
                break
            # Look for STT header
            if vals and str(vals[0]).strip() in ('STT', 'S\nT\nT'):
                header_row = row_idx
                for i, v in enumerate(vals):
                    if v:
                        v_clean = str(v).strip().replace('\n', ' ')
                        col_map[v_clean.upper()] = i
        
        if not data_start_row:
            if header_row:
                data_start_row = header_row + 4  # Skip sub-headers
            else:
                continue
        
        # Determine column indices based on common patterns
        # Columns: STT(0), Tên TS(1), Mã số(2), Năm SĐ(3), SL sổ sách(4), SL thực tế(5), 
        #          Chênh lệch(6), Đơn giá(7), Thành tiền(8), BP quản lý(9), 
        #          Nơi sử dụng(10), Tình trạng(11), Nơi chuyển đến(12), Ghi chú(13)
        
        for row in ws.iter_rows(min_row=data_start_row, max_row=ws.max_row, values_only=False):
            vals = [cell.value for cell in row]
            
            # Skip empty rows, header rows, and summary rows
            stt = vals[0] if len(vals) > 0 else None
            name = vals[1] if len(vals) > 1 else None
            asset_code = vals[2] if len(vals) > 2 else None
            
            if not name or not asset_code:
                continue
            
            # Skip if it's a section header (no STT number)
            if stt is None and asset_code is None:
                continue
            
            # Skip sub-items (no STT, no asset code)
            name = clean_text(name)
            asset_code = clean_text(asset_code)
            
            if not asset_code or not name:
                continue
            
            # Clean asset code (remove newlines)
            asset_code = asset_code.replace('\n', '').replace('\r', '').strip()
            
            # Check if already exists
            cursor = conn.execute("SELECT id FROM Asset WHERE assetCode = ?", (asset_code,))
            if cursor.fetchone():
                skipped += 1
                continue
            
            year_in_use = safe_int(vals[3] if len(vals) > 3 else None)
            book_qty = safe_int(vals[4] if len(vals) > 4 else None) or 1
            actual_qty = safe_int(vals[5] if len(vals) > 5 else None) or book_qty
            qty_diff = safe_int(vals[6] if len(vals) > 6 else None) or 0
            unit_price = safe_float(vals[7] if len(vals) > 7 else None)
            total_price = safe_float(vals[8] if len(vals) > 8 else None)
            
            bp_quanly = clean_text(vals[9] if len(vals) > 9 else None)
            noi_su_dung = clean_text(vals[10] if len(vals) > 10 else None)
            tinh_trang = clean_text(vals[11] if len(vals) > 11 else None)
            noi_chuyen_den = clean_text(vals[12] if len(vals) > 12 else None)
            ghi_chu = clean_text(vals[13] if len(vals) > 13 else None)
            
            status = get_status(tinh_trang)
            qr_code = str(uuid.uuid4())[:12]
            
            # Determine location
            location = 'Cơ sở 1'
            if noi_chuyen_den and ('cơ sở 2' in noi_chuyen_den.lower() or 'cs2' in noi_chuyen_den.lower()):
                location = 'Cơ sở 2'
            
            try:
                conn.execute("""
                    INSERT INTO Asset (assetCode, name, categoryId, departmentId, location, locationDetail,
                        assignedTo, yearInUse, originalPrice, currentValue, status,
                        bookQuantity, actualQuantity, quantityDifference, source, note, qrCode,
                        createdAt, updatedAt)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    asset_code, name, category_id, dept_id, location, noi_su_dung,
                    None, year_in_use, unit_price or total_price, None, status,
                    book_qty, actual_qty, qty_diff, bp_quanly, 
                    '; '.join(filter(None, [ghi_chu, noi_chuyen_den])) or None,
                    qr_code, int(datetime.now().timestamp() * 1000), int(datetime.now().timestamp() * 1000)
                ))
                imported += 1
            except Exception as e:
                print(f"  Error importing {asset_code}: {e}")
                skipped += 1
    
    conn.commit()
    print(f"TBVP Main: Imported {imported}, Skipped {skipped}")

def import_cntt(conn):
    """Import thiết bị CNTT từ các file trong thư mục CNTT"""
    folder = r'f:\QLTS\TS\CNTT'
    if not os.path.exists(folder):
        print(f"Folder not found: {folder}")
        return
    
    category_id = get_category_id(conn, 'CNTT')
    imported = 0
    skipped = 0
    
    # File to department mapping based on file names
    file_dept_map = {
        'PKĐK': 'PKDK', 'PKDK': 'PKDK',
        'BKLN': 'PCBKLN', 'KLN': 'PCBKLN',
        'Dược': 'DVTYT', 'DUOC': 'DVTYT',
        'HIV': 'HIV',
        'BTN': 'PCBTN',
        'TCKT': 'TCKT',
        'TCHC': 'TCHC',
        'KDYTQT': 'KDYTQT',
        'SKSS': 'SKSS',
        'SKMT': 'SKMT',
        'DD': 'DD', 'Dinh dưỡng': 'DD',
        'KHNV': 'KHNV',
        'TTGDSK': 'TTGDSK',
        'KSTCT': 'KSTCT', 'KST': 'KSTCT',
        'BNN': 'BNN',
        'XN': 'XN',
    }
    
    for fn in sorted(os.listdir(folder)):
        if not fn.endswith('.xlsx'):
            continue
        
        # Find department from filename
        dept_code = None
        for key, code in file_dept_map.items():
            if key.lower() in fn.lower():
                dept_code = code
                break
        
        if not dept_code:
            print(f"  Cannot determine department for: {fn}")
            continue
        
        cursor = conn.execute("SELECT id FROM Department WHERE code = ?", (dept_code,))
        dept_row = cursor.fetchone()
        if not dept_row:
            continue
        dept_id = dept_row[0]
        
        filepath = os.path.join(folder, fn)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"  Error opening {fn}: {e}")
            continue
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row <= 5 or ws.max_column <= 3:
                continue
            
            # Find data start row
            data_start = None
            for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=False), 1):
                vals = [cell.value for cell in row]
                if vals and vals[0] == 'A' and vals[1] == 'B':
                    data_start = row_idx + 1
                    break
            
            if not data_start:
                continue
            
            for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False):
                vals = [cell.value for cell in row]
                
                stt = vals[0] if len(vals) > 0 else None
                name = vals[1] if len(vals) > 1 else None
                
                if not name or stt is None:
                    continue
                
                name = clean_text(name)
                if not name:
                    continue
                
                # CNTT files may have different column layouts
                # Try to find asset code (col 2 or 3)
                asset_code = None
                year_col = None
                
                if len(vals) > 2:
                    code_val = clean_text(vals[2])
                    if code_val and not code_val.replace(' ', '').isdigit():
                        asset_code = code_val
                
                if len(vals) > 3 and not asset_code:
                    code_val = clean_text(vals[3])
                    if code_val and not code_val.replace(' ', '').isdigit():
                        asset_code = code_val
                
                # Generate asset code if not found
                if not asset_code:
                    asset_code = f"CNTT-{dept_code}-{str(uuid.uuid4())[:6]}"
                
                asset_code = asset_code.replace('\n', '').replace('\r', '').strip()
                
                # Check duplicate
                cursor = conn.execute("SELECT id FROM Asset WHERE assetCode = ?", (asset_code,))
                if cursor.fetchone():
                    skipped += 1
                    continue
                
                # Extract year
                year_in_use = None
                for i in range(2, min(6, len(vals))):
                    y = safe_int(vals[i])
                    if y and 1990 <= y <= 2030:
                        year_in_use = y
                        break
                
                qr_code = str(uuid.uuid4())[:12]
                
                # Get status and other fields from later columns
                status = 'DANG_SU_DUNG'
                note_parts = []
                noi_su_dung = None
                
                # Look for status in known column positions
                for i in range(len(vals) - 5, len(vals)):
                    if i >= 0 and vals[i]:
                        text = str(vals[i]).strip()
                        if any(s in text.lower() for s in ['sử dụng', 'hỏng', 'thanh lý', 'bảo trì']):
                            status = get_status(text)
                            break
                
                try:
                    conn.execute("""
                        INSERT INTO Asset (assetCode, name, categoryId, departmentId, location,
                            yearInUse, status, bookQuantity, actualQuantity, quantityDifference,
                            note, qrCode, createdAt, updatedAt)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        asset_code, name, category_id, dept_id, 'Cơ sở 1',
                        year_in_use, status, 1, 1, 0,
                        None, qr_code, int(datetime.now().timestamp() * 1000), int(datetime.now().timestamp() * 1000)
                    ))
                    imported += 1
                except Exception as e:
                    skipped += 1
    
    conn.commit()
    print(f"CNTT: Imported {imported}, Skipped {skipped}")

def import_tbvp_inventory(conn):
    """Import dữ liệu kiểm kê TBVP từ các file khoa/phòng"""
    folder = r'f:\QLTS\TS\TBVP Kiem ke tai san hoan thien'
    if not os.path.exists(folder):
        print(f"Folder not found: {folder}")
        return
    
    category_id = get_category_id(conn, 'TBVP')
    imported = 0
    skipped = 0
    
    file_dept_map = {
        'BNN': 'BNN', 'SKSS': 'SKSS', 'PCBKLN': 'PCBKLN',
        'KDYTQT': 'KDYTQT', 'PCBTN': 'PCBTN', 'KSTCT': 'KSTCT',
        'TTGDSK': 'TTGDSK', 'DD': 'DD', 'HIV': 'HIV',
        'PKDK': 'PKDK', 'PKĐK': 'PKDK', 'KHNV': 'KHNV',
        'SKMT': 'SKMT', 'XN': 'XN', 'DUOC': 'DVTYT', 'Dược': 'DVTYT',
        'TCHC': 'TCHC', 'TCKT': 'TCKT',
    }
    
    for fn in sorted(os.listdir(folder)):
        if not fn.endswith('.xlsx'):
            continue
        
        dept_code = None
        fn_upper = fn.upper()
        for key, code in file_dept_map.items():
            if key.upper() in fn_upper:
                dept_code = code
                break
        
        if not dept_code:
            print(f"  Cannot determine department for: {fn}")
            continue
        
        cursor = conn.execute("SELECT id FROM Department WHERE code = ?", (dept_code,))
        dept_row = cursor.fetchone()
        if not dept_row:
            continue
        dept_id = dept_row[0]
        
        filepath = os.path.join(folder, fn)
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
        except Exception as e:
            print(f"  Error opening {fn}: {e}")
            continue
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            if ws.max_row <= 5 or ws.max_column <= 3:
                continue
            
            # Find header/data start
            data_start = None
            for row_idx in range(1, min(30, ws.max_row + 1)):
                row = [ws.cell(row=row_idx, column=c).value for c in range(1, min(ws.max_column + 1, 50))]
                if row and row[0] == 'A' and len(row) > 1 and row[1] == 'B':
                    data_start = row_idx + 1
                    break
            
            if not data_start:
                continue
            
            for row_idx in range(data_start, ws.max_row + 1):
                vals = [ws.cell(row=row_idx, column=c).value for c in range(1, min(ws.max_column + 1, 50))]
                
                stt = vals[0] if len(vals) > 0 else None
                name = vals[1] if len(vals) > 1 else None
                asset_code = vals[2] if len(vals) > 2 else None
                
                if not name:
                    continue
                
                name = clean_text(name)
                if not name:
                    continue
                
                if asset_code:
                    asset_code = str(asset_code).replace('\n', '').replace('\r', '').strip()
                
                if not asset_code:
                    continue
                
                # Check duplicate
                cursor = conn.execute("SELECT id FROM Asset WHERE assetCode = ?", (asset_code,))
                if cursor.fetchone():
                    skipped += 1
                    continue
                
                year_in_use = safe_int(vals[3] if len(vals) > 3 else None)
                if year_in_use and (year_in_use < 1990 or year_in_use > 2030):
                    year_in_use = None
                
                book_qty = safe_int(vals[4] if len(vals) > 4 else None) or 1
                actual_qty = safe_int(vals[5] if len(vals) > 5 else None) or book_qty
                qty_diff = safe_int(vals[6] if len(vals) > 6 else None) or 0
                unit_price = safe_float(vals[7] if len(vals) > 7 else None)
                
                tinh_trang = clean_text(vals[11] if len(vals) > 11 else None)
                status = get_status(tinh_trang)
                
                ghi_chu = clean_text(vals[13] if len(vals) > 13 else None)
                noi_su_dung = clean_text(vals[10] if len(vals) > 10 else None)
                
                qr_code = str(uuid.uuid4())[:12]
                
                try:
                    conn.execute("""
                        INSERT INTO Asset (assetCode, name, categoryId, departmentId, location,
                            locationDetail, yearInUse, originalPrice, status,
                            bookQuantity, actualQuantity, quantityDifference,
                            note, qrCode, createdAt, updatedAt)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        asset_code, name, category_id, dept_id, 'Cơ sở 1',
                        noi_su_dung, year_in_use, unit_price, status,
                        book_qty, actual_qty, qty_diff,
                        ghi_chu, qr_code, int(datetime.now().timestamp() * 1000), int(datetime.now().timestamp() * 1000)
                    ))
                    imported += 1
                except Exception as e:
                    skipped += 1
    
    conn.commit()
    print(f"TBVP Inventory: Imported {imported}, Skipped {skipped}")

def main():
    if not os.path.exists(DB_PATH):
        print(f"Database not found at {DB_PATH}")
        print("Please run 'npx prisma migrate dev' first in the server directory.")
        return
    
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    
    # Check if departments exist
    cursor = conn.execute("SELECT COUNT(*) FROM Department")
    dept_count = cursor.fetchone()[0]
    if dept_count == 0:
        print("No departments found. Please run seed first: npx ts-node prisma/seed.ts")
        conn.close()
        return
    
    print("="*60)
    print("IMPORT DỮ LIỆU TÀI SẢN CDC ĐÀ NẴNG")
    print("="*60)
    
    # Count existing assets
    cursor = conn.execute("SELECT COUNT(*) FROM Asset")
    existing = cursor.fetchone()[0]
    print(f"Existing assets in DB: {existing}")
    
    print("\n--- Importing TBVP (Tổng hợp) ---")
    import_tbvp_main(conn)
    
    print("\n--- Importing TBVP (Kiểm kê chi tiết) ---")
    import_tbvp_inventory(conn)
    
    print("\n--- Importing CNTT ---")
    import_cntt(conn)
    
    # Final count
    cursor = conn.execute("SELECT COUNT(*) FROM Asset")
    total = cursor.fetchone()[0]
    print(f"\n{'='*60}")
    print(f"TOTAL ASSETS IN DATABASE: {total}")
    
    # Summary by category
    cursor = conn.execute("""
        SELECT ac.name, COUNT(a.id) 
        FROM Asset a 
        JOIN AssetCategory ac ON a.categoryId = ac.id 
        GROUP BY ac.name
    """)
    print("\nBy Category:")
    for row in cursor:
        print(f"  {row[0]}: {row[1]}")
    
    # Summary by department
    cursor = conn.execute("""
        SELECT d.code, d.name, COUNT(a.id) 
        FROM Asset a 
        JOIN Department d ON a.departmentId = d.id 
        GROUP BY d.code 
        ORDER BY COUNT(a.id) DESC
    """)
    print("\nBy Department:")
    for row in cursor:
        print(f"  {row[0]} ({row[1]}): {row[2]}")
    
    # Summary by status
    cursor = conn.execute("""
        SELECT status, COUNT(*) 
        FROM Asset 
        GROUP BY status 
        ORDER BY COUNT(*) DESC
    """)
    print("\nBy Status:")
    for row in cursor:
        print(f"  {row[0]}: {row[1]}")
    
    conn.close()
    print(f"\n{'='*60}")
    print("Import completed!")

if __name__ == '__main__':
    main()
